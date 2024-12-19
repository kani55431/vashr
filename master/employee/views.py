from django.shortcuts import render, get_object_or_404, redirect
from .models import Employee, Wages
from .forms import EmployeeForm, WagesForm, WagesEditForm, EmployeeEditForm
from django.core.exceptions import ValidationError 
from django.contrib import messages


from django.shortcuts import render
from django.db.models import Count, Q
from .models import Employee
from django.utils import timezone
from datetime import date

def hr_dashboard(request):
    # Count active employees
    total_active_employees = Employee.objects.filter(status='active').count()

    # Count employees with PF and without PF
    pf_count = Employee.objects.filter(status='active', pf_esi_applicable='yes').count()
    no_pf_count = Employee.objects.filter(status='active', pf_esi_applicable='no').count()

    # Count male and female employees
    male_count = Employee.objects.filter(status='active', gender='male').count()
    female_count = Employee.objects.filter(status='active', gender='female').count()
    yesterday = timezone.now().date() - timezone.timedelta(days=1)
    
    # Query present employees for yesterday
    present_employees = Attendance.objects.filter(date=yesterday, absent='P')

    # Calculate total wages
    total_amount = 0
    departments = []
    wages = []
    
    today = date.today()  # Get today's date

    for attendance in present_employees:
        wage = Wages.objects.filter(employee=attendance.employee).first()  # Get the wage record
        if wage and wage.start_date <= today <= (wage.end_date or today):  # Check if today is within the wage period
            total_amount += wage.per_day
            departments.append(attendance.department)
            wages.append(wage.per_day)
    # Get department-wise employee count
    department_stats = Employee.objects.filter(status='active').values('department').annotate(active_employees=Count('id'))
    
    # Get total departments
    total_departments = department_stats.count()

    # Group active employees by department
    departments = Department.objects.prefetch_related('employee_set').all()  # Assuming Employee has a ForeignKey to Department
    department_data = {department.name: department.employee_set.filter(status='active').count() for department in departments}

    context = {
        'total_active_employees': total_active_employees,
        'pf_count': pf_count,
        'no_pf_count': no_pf_count,
        'male_count': male_count,
        'female_count': female_count,
         'yesterday': yesterday,
        'total_amount': total_amount,
        'departments': departments,
        'total_departments':total_departments,
        'wages': wages,
         'department_data': department_data  # Add department data to context
        
    }
    return render(request, 'body/hr_dashboard.html', context)




def employee_idcard(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    return render(request, 'employee/employee_idcard.html', {'employee': employee})

def employee_list(request):
    employees = Employee.objects.filter(status='active')
    return render(request, 'employee/employee_list.html', {'employees': employees})


def employee_detail(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    return render(request, 'employee/employee_detail.html', {'employee': employee})

def employee_create(request):
    if request.method == 'POST':
        form = EmployeeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('employee_list')
    else:
        form = EmployeeForm()
    return render(request, 'employee/employee_form.html', {'form': form})


from django.shortcuts import render, get_object_or_404, redirect
from .models import Employee
from .forms import EmployeeEditForm

def employee_update(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        # Include request.FILES to handle file uploads
        form = EmployeeEditForm(request.POST, request.FILES, instance=employee)
        if form.is_valid():
            form.save()
            return redirect('employee_list')
    else:
        form = EmployeeEditForm(instance=employee)  # Passing the employee instance

    return render(request, 'employee/employee_edit_form.html', {'form': form})


def employee_delete(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        employee.delete()
        return redirect('employee_list')
    return render(request, 'employee/employee_confirm_delete.html', {'object': employee})
from django.shortcuts import render
from django.views.generic import ListView
from .models import Employee

class InactiveEmployeeListView(ListView):
    model = Employee
    template_name = 'employee/inactive_employees.html'
    context_object_name = 'inactive_employees'

    def get_queryset(self):
        return Employee.objects.filter(status='inactive')
    


#######################################
from django.http import HttpResponse
from django.views.generic import View
import pandas as pd
from .models import Employee

class EmployeeExportView(View):
    def get(self, request):
        # Query all employees
        employees = Employee.objects.all()

        # Prepare data for DataFrame
        data = {
            'Employee Code': [emp.employee_code for emp in employees],
            'Employee Name': [emp.employee_name for emp in employees],
            'Employee Mobile': [emp.employee_mobile for emp in employees],
            'Emergency Mobile': [emp.emergency_mobile for emp in employees],
            'Father or Husband Name': [emp.father_or_husband_name for emp in employees],
            'Address Line 1': [emp.address_1 for emp in employees],
            'Street': [emp.street for emp in employees],
            'City': [emp.city for emp in employees],
            'District': [emp.district for emp in employees],
            'State': [emp.state for emp in employees],
            'Pincode': [emp.pincode for emp in employees],
            'Date of Birth': [emp.date_of_birth.strftime('%Y-%m-%d') if emp.date_of_birth else '' for emp in employees],
            'Age': [emp.age for emp in employees],
            'Qualification': [emp.qualification for emp in employees],
            'Marital Status': [emp.marital_status for emp in employees],
            'Blood Group': [emp.blood_group for emp in employees],
            'Gender': [emp.gender for emp in employees],
            'Department': [emp.department for emp in employees],
            'Category': [emp.category for emp in employees],
            'Designation': [emp.designation for emp in employees],
            'Weekly Off': [emp.weekly_off for emp in employees],
            'Date of Joining': [emp.date_of_joining.strftime('%Y-%m-%d %H:%M:%S') if emp.date_of_joining else '' for emp in employees],
            'Date of Re-joining': [emp.date_of_re_joining.strftime('%Y-%m-%d %H:%M:%S') if emp.date_of_re_joining else '' for emp in employees],
            'Date of Leaving': [emp.date_of_leaving.strftime('%Y-%m-%d %H:%M:%S') if emp.date_of_leaving else '' for emp in employees],
            'PF/ESI Applicable': [emp.pf_esi_applicable for emp in employees],
            'PF Date of Joining': [emp.pf_date_of_joining.strftime('%Y-%m-%d %H:%M:%S') if emp.pf_date_of_joining else '' for emp in employees],
            'PF Number': [emp.pf_no for emp in employees],
            'ESI Number': [emp.esi_no for emp in employees],
            'Bank Name': [emp.bank_name for emp in employees],
            'Account Holder': [emp.account_holder for emp in employees],
            'Account Number': [emp.account_no for emp in employees],
            'IFSC Code': [emp.ifsc_code for emp in employees],
            'Branch': [emp.branch for emp in employees],
            'Nominee Name': [emp.nominee_name for emp in employees],
            'Nominee Relationship': [emp.nominee_relationship for emp in employees],
            'Nominee Age': [emp.nominee_age for emp in employees],
            'Employee Photo': [emp.employee_photo.url if emp.employee_photo else '' for emp in employees],
            'Aadhar Number': [emp.aadhar_number for emp in employees],
            'Migrant Worker': [emp.migrant_worker for emp in employees],
            'Working Mode': [emp.working_mode for emp in employees],
            'Status': [emp.status for emp in employees],
        
        
        }

        # Create DataFrame
        df = pd.DataFrame(data)

        # Create Excel file
        excel_file = 'employees.xlsx'
        df.to_excel(excel_file, index=False)

        # Prepare response
        response = HttpResponse(open(excel_file, 'rb'), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="employees.xlsx"'

        return response

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from .models import Designation, Category, Department
from .forms import DesignationForm, CategoryForm, DepartmentForm

# Designation Views
def designation_list(request):
    designations = Designation.objects.all()
    return render(request, 'employee/department/designation_list.html', {'designations': designations})

def designation_create(request):
    if request.method == 'POST':
        form = DesignationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('designation_list')
    else:
        form = DesignationForm()
    return render(request, 'employee/department/designation_form.html', {'form': form})

def designation_edit(request, pk):
    designation = get_object_or_404(Designation, pk=pk)
    if request.method == 'POST':
        form = DesignationForm(request.POST, instance=designation)
        if form.is_valid():
            form.save()
            return redirect('designation_list')
    else:
        form = DesignationForm(instance=designation)
    return render(request, 'employee/department/designation_form.html', {'form': form})

def designation_delete(request, pk):
    designation = get_object_or_404(Designation, pk=pk)
    if request.method == 'POST':
        designation.delete()
        return redirect('designation_list')
    return render(request, 'employee/department/designation_confirm_delete.html', {'designation': designation})

# Category Views
def category_list(request):
    categories = Category.objects.all()
    return render(request, 'employee/department/category_list.html', {'categories': categories})

def category_create(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('category_list')
    else:
        form = CategoryForm()
    return render(request, 'employee/department/category_form.html', {'form': form})

def category_edit(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            return redirect('category_list')
    else:
        form = CategoryForm(instance=category)
    return render(request, 'employee/department/category_form.html', {'form': form})

def category_delete(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        category.delete()
        return redirect('category_list')
    return render(request, 'employee/department/category_confirm_delete.html', {'category': category})

# Department Views
def department_list(request):
    departments = Department.objects.all()
    return render(request, 'employee/department/department_list.html', {'departments': departments})

def department_create(request):
    if request.method == 'POST':
        form = DepartmentForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('department_list')
    else:
        form = DepartmentForm()
    return render(request, 'employee/department/department_form.html', {'form': form})

def department_edit(request, pk):
    department = get_object_or_404(Department, pk=pk)
    if request.method == 'POST':
        form = DepartmentForm(request.POST, instance=department)
        if form.is_valid():
            form.save()
            return redirect('department_list')
    else:
        form = DepartmentForm(instance=department)
    return render(request, 'employee/department/department_form.html', {'form': form})

def department_delete(request, pk):
    department = get_object_or_404(Department, pk=pk)
    if request.method == 'POST':
        department.delete()
        return redirect('department_list')
    return render(request, 'employee/department/department_confirm_delete.html', {'department': department})


def master_creation(request):
    return render(request, 'employee/department/master_creation.html',)

from django.shortcuts import render
from django.views import View
from .forms import EmployeeUploadForm
from .models import Employee
import pandas as pd
from django.views import View
from django.shortcuts import render
from django.core.exceptions import ObjectDoesNotExist
from .models import Employee, Category, Department
import pandas as pd
from django.views import View
from django.shortcuts import render
from django.db import IntegrityError  # Add this import
from .models import Employee, Category, Department
import pandas as pd
from django.views import View
from django.shortcuts import render
from django.db import IntegrityError
from .models import Employee, Category, Department
import pandas as pd
from django.views import View
from django.shortcuts import render
from django.db import IntegrityError
from .models import Employee, Category, Department
import pandas as pd

from django.views import View
from django.shortcuts import render
from django.db import IntegrityError
from .models import Employee, Category, Department
import pandas as pd
import pandas as pd
from django.shortcuts import render
from django.views import View
from django.db import IntegrityError
import pytz
from .models import Category, Department, Designation, Employee
from .forms import EmployeeUploadForm  # Ensure you have this form defined

import pandas as pd
import numpy as np
import pytz
from django.shortcuts import render
from django.views import View
from django.db import IntegrityError
from .models import Employee, Department, Category, Designation
from .forms import EmployeeUploadForm
import pandas as pd
import numpy as np
import pytz
from django.shortcuts import render
from django.views import View
from django.db import IntegrityError
from .models import Employee, Department, Category, Designation
from .forms import EmployeeUploadForm
import pandas as pd
import pytz

from django.shortcuts import render
from django.views import View
from django.db import IntegrityError
from .models import Employee, Department, Category, Designation
from .forms import EmployeeUploadForm
import pandas as pd
import pytz

from django.shortcuts import render
from django.views import View
from django.db import IntegrityError
from .models import Employee, Department, Category, Designation
from .forms import EmployeeUploadForm
import pandas as pd
import pytz
from django.utils.dateparse import parse_date

import pandas as pd
from django.utils.dateparse import parse_date  # Add this import for parsing date strings
import numpy as np
from django.shortcuts import render
from django.db import IntegrityError
import pandas as pd
from .models import Employee, Category, Department, Designation
from .forms import EmployeeUploadForm


def upload_employee_data(request):
    if request.method == 'POST':
        form = EmployeeUploadForm(request.POST, request.FILES)

        if form.is_valid():
            excel_file = request.FILES['file']

            # Variables for counting success and errors
            success_count = 0
            error_count = 0
            error_messages = []

            try:
                # Load the Excel file using pandas
                df = pd.read_excel(excel_file)

                # Iterate over each row in the Excel sheet
                for index, row in df.iterrows():
                    try:
                        # Normalize and validate Gender
                        gender_raw = str(row.get('Gender', '')).strip().lower() if pd.notna(row.get('Gender')) else ''
                        gender = 'male' if gender_raw in ['male', 'm'] else 'female' if gender_raw in ['female', 'f'] else None

                        # Normalize and validate Weekly Off
                        weekly_off_raw = str(row.get('Weekly Off', '')).strip().lower() if pd.notna(row.get('Weekly Off')) else ''
                        weekly_off_options = {
                            'sunday': 'Sunday', 'monday': 'Monday', 'tuesday': 'Tuesday',
                            'wednesday': 'Wednesday', 'thursday': 'Thursday',
                            'friday': 'Friday', 'saturday': 'Saturday'
                        }
                        weekly_off = weekly_off_options.get(weekly_off_raw, None)

                        # Normalize and validate Marital Status
                        marital_status_raw = str(row.get('Marital Status', '')).strip().lower() if pd.notna(row.get('Marital Status')) else ''
                        marital_status_options = {
                            'single': 'Single', 'married': 'Married', 'widowed': 'Widowed', 'divorced': 'Divorced'
                        }
                        marital_status = marital_status_options.get(marital_status_raw, None)

                        # Retrieve or create related objects (Category, Department, Designation)
                        category_name = str(row.get('Category', '')).strip() if pd.notna(row.get('Category')) else None
                        category, _ = Category.objects.get_or_create(name=category_name) if category_name else (None, False)

                        department_name = str(row.get('Department', '')).strip() if pd.notna(row.get('Department')) else None
                        department, created_department = Department.objects.get_or_create(
                            name=department_name,
                            defaults={'category': category}
                        ) if department_name else (None, False)

                        if department and category and department.category != category:
                            department.category = category
                            department.save()

                        designation_name = str(row.get('Designation', '')).strip() if pd.notna(row.get('Designation')) else None
                        designation = Designation.objects.get_or_create(name=designation_name)[0] if designation_name else None

                        # Extract and validate date fields
                        date_of_birth = pd.to_datetime(row.get('Date of Birth'), errors='coerce') if pd.notnull(row.get('Date of Birth')) else None
                        date_of_joining = pd.to_datetime(row.get('Date of Joining'), errors='coerce') if pd.notnull(row.get('Date of Joining')) else None
                        date_of_leaving = pd.to_datetime(row.get('Date of Leaving'), errors='coerce') if pd.notnull(row.get('Date of Leaving')) else None
                        pf_date_of_joining = pd.to_datetime(row.get('PF Date of Joining'), errors='coerce') if pd.notnull(row.get('PF Date of Joining')) else None

                        # Retrieve or create the Employee
                        employee_code = str(row.get('Employee Code', '')).strip() if pd.notna(row.get('Employee Code')) else None
                        if not employee_code:
                            error_count += 1
                            error_messages.append(f"Missing Employee Code in row {index + 1}")
                            continue

                        employee, created = Employee.objects.get_or_create(
                            employee_code=employee_code,
                            defaults={
                                'employee_name': row.get('Employee Name', ''),
                                'employee_mobile': row.get('Employee Mobile', ''),
                                'emergency_mobile': row.get('Emergency Mobile', ''),
                                'father_or_husband_name': row.get('Father or Husband Name', ''),
                                'address_1': row.get('Address Line 1', ''),
                                'street': row.get('Street', ''),
                                'city': row.get('City', ''),
                                'district': row.get('District', ''),
                                'state': row.get('State', ''),
                                'pincode': row.get('Pincode', ''),
                                'date_of_birth': date_of_birth,
                                'age': int(row.get('Age', 0)) if pd.notnull(row.get('Age')) else None,
                                'qualification': row.get('Qualification', ''),
                                'marital_status': marital_status,
                                'blood_group': row.get('Blood Group', ''),
                                'gender': gender,
                                'department': department,
                                'category': category,
                                'designation': designation,
                                'date_of_joining': date_of_joining,
                                'date_of_leaving': date_of_leaving,
                                'pf_esi_applicable': row.get('PF/ESI Applicable', ''),
                                'pf_date_of_joining': pf_date_of_joining,
                                'pf_no': row.get('PF Number', ''),
                                'esi_no': row.get('ESI Number', ''),
                                'bank_name': row.get('Bank Name', ''),
                                'account_holder': row.get('Account Holder', ''),
                                'account_no': row.get('Account Number', ''),
                                'ifsc_code': row.get('IFSC Code', ''),
                                'branch': row.get('Branch', ''),
                                'nominee_name': row.get('Nominee Name', ''),
                                'nominee_relationship': row.get('Nominee Relationship', ''),
                                'nominee_age': int(row.get('Nominee Age', 0)) if pd.notnull(row.get('Nominee Age')) else None,
                                'food_expense': row.get('Food Expense', ''),
                                'migrant_worker_specific': row.get('Migrant Worker Specific', ''),
                                'migrant_worker': row.get('Migrant Worker', ''),
                                'working_mode': row.get('Working Mode', ''),
                                'weekly_off': weekly_off,
                                'status': row.get('Status', ''),
                            }
                        )

                        # Update fields if employee already exists
                        if not created:
                            for field in [
                                'employee_name', 'employee_mobile', 'emergency_mobile',
                                'father_or_husband_name', 'address_1', 'street', 'city', 'district',
                                'state', 'pincode', 'qualification', 'marital_status', 'blood_group',
                                'weekly_off', 'status'
                            ]:
                                setattr(employee, field, row.get(field, getattr(employee, field)))

                            employee.date_of_birth = date_of_birth or employee.date_of_birth
                            employee.date_of_joining = date_of_joining or employee.date_of_joining
                            employee.date_of_leaving = date_of_leaving or employee.date_of_leaving
                            employee.pf_date_of_joining = pf_date_of_joining or employee.pf_date_of_joining
                            employee.department = department or employee.department
                            employee.category = category or employee.category
                            employee.designation = designation or employee.designation
                            employee.gender = gender or employee.gender
                            employee.save()

                        success_count += 1

                    except IntegrityError as e:
                        error_count += 1
                        error_messages.append(f"IntegrityError with Employee Code '{employee_code}': {str(e)}")
                    except Exception as e:
                        error_count += 1
                        error_messages.append(f"Unexpected error with Employee Code '{employee_code}': {str(e)}")

                # Prepare success message
                success_message = f"{success_count} employees successfully created/updated."
                return render(request, 'fileupload/employeeexcelupload.html', {
                    'form': form,
                    'success_message': success_message,
                    'error_count': error_count,
                    'error_messages': error_messages
                })

            except Exception as e:
                # Handle file processing errors
                error_messages.append(f"Error processing file: {str(e)}")
                return render(request, 'fileupload/employeeexcelupload.html', {
                    'form': form,
                    'error_messages': error_messages
                })

        else:
            # Handle form validation errors
            error_messages = [f"Form errors: {form.errors}"]
            return render(request, 'fileupload/employeeexcelupload.html', {
                'form': form,
                'error_messages': error_messages
            })

    else:
        # For GET requests, render an empty form
        form = EmployeeUploadForm()
        return render(request, 'fileupload/employeeexcelupload.html', {'form': form})


from django.shortcuts import render, redirect, get_object_or_404
from .forms import WagesForm
from .models import Wages

def round_up(value):
    return round(value) if value % 1 < 0.5 else round(value) + 0

def wages_create_view(request):
    if request.method == 'POST':
        form = WagesForm(request.POST)
        if form.is_valid():
            wages = form.save(commit=False)
            wages.basic_amount = round_up((wages.per_day * wages.basic_precentage) / 100)
            wages.da_amount = round_up((wages.per_day * wages.da_precentage) / 100)
            wages.hra_amount = round_up((wages.per_day * wages.hra_precentage) / 100)
            wages.save()
            return redirect('wages_list')
    else:
        form = WagesForm()
    return render(request, 'employee/employee_wages/wages_form.html', {'form': form})

def wages_update_view(request, wage_id):
    wages = get_object_or_404(Wages, id=wage_id)
    if request.method == 'POST':
        form = WagesEditForm(request.POST, instance=wages)
        if form.is_valid():
            wages = form.save(commit=False)
            wages.basic_amount = round_up((wages.per_day * wages.basic_precentage) / 100)
            wages.da_amount = round_up((wages.per_day * wages.da_precentage) / 100)
            wages.hra_amount = round_up((wages.per_day * wages.hra_precentage) / 100)
            wages.other_allowance = round_up((wages.per_day * wages.other_allowance_precentage) / 100)
            

            wages.save()
            return redirect('wages_list')
    else:
        form = WagesEditForm(instance=wages)
    return render(request, 'employee/employee_wages/wages_form.html', {'form': form})





def wages_delete_view(request, wage_id):
    wages = get_object_or_404(Wages, id=wage_id)
    if request.method == 'POST':
        wages.delete()
        messages.success(request, 'Wages record deleted successfully.')
        return redirect('wages_list')  # Redirect to the list view
    return render(request, 'employee/employee_wages/wages_confirm_delete.html', {'wages': wages})



def wages_reports(request):
    return render(request, 'employee/employee_wages/wages_reports.html')


def wages_list_view(request):
    wages_list = Wages.objects.select_related('employee').all()
    return render(request, 'employee/employee_wages/wages_list.html', {'wages_list': wages_list})


# views.py
import pandas as pd
from django.shortcuts import render, redirect
from .forms import WagesUploadForm
from .models import Wages, Employee
# views.py
import pandas as pd
from django.shortcuts import render, redirect
from .forms import WagesUploadForm
from .models import Wages, Employee

def upload_wages(request):
    if request.method == "POST":
        form = WagesUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            try:
                # Read the Excel file
                df = pd.read_excel(file)

                errors = []  # To collect any errors during processing

                # Iterate through the rows in the DataFrame
                for index, row in df.iterrows():
                    employee_code = row['employee_code']  # Get the employee_code from the Excel

                    # Check if the Employee exists with the given employee_code
                    if Employee.objects.filter(employee_code=employee_code).exists():
                        employee = Employee.objects.get(employee_code=employee_code)

                        # Extract values from the row
                        per_day = row['per_day']
                        basic_precentage = row['basic_precentage']
                        da_precentage = row['da_precentage']
                        hra_precentage = row['hra_precentage']
                        other_allowance_precentage = row['other_allowance_precentage']

                        # Calculate the amounts
                        basic_amount = (basic_precentage / 100) * per_day
                        da_amount = (da_precentage / 100) * per_day
                        hra_amount = (hra_precentage / 100) * per_day
                        other_allowance = (other_allowance_precentage / 100) * per_day

                        # Create a new Wages entry
                        Wages.objects.create(
                            employee=employee,
                            per_day=per_day,
                            basic_precentage=basic_precentage,
                            da_precentage=da_precentage,
                            hra_precentage=hra_precentage,
                            other_allowance_precentage=other_allowance_precentage,
                            basic_amount=basic_amount,
                            da_amount=da_amount,
                            hra_amount=hra_amount,
                            other_allowance=other_allowance,
                            rent=row['rent'],
                            start_date=row['start_date'],
                            end_date=row['end_date']
                        )
                    else:
                        errors.append(f"Employee code {employee_code} does not exist at row {index + 1}.")
                
                if errors:
                    # If there are errors, re-render the form with errors
                    for error in errors:
                        form.add_error(None, error)
                    return render(request, 'employee/employee_wages/upload_wages.html', {'form': form})

                return redirect('wages_list')  # Redirect to a success page
            except Exception as e:
                # Handle any other exceptions
                form.add_error(None, f"There was an error processing the file: {str(e)}")
    else:
        form = WagesUploadForm()

    return render(request, 'employee/employee_wages/upload_wages.html', {'form': form})


from django.shortcuts import render, redirect
from django.http import JsonResponse
from .forms import IncentiveForm, IncentiveAmountForm
from .models import Employee, Incentive

def incentive_create_view(request):
    if request.method == 'POST':
        form = IncentiveForm(request.POST)
        if form.is_valid():
            date = form.cleaned_data['date']
            table_data = request.POST.get('table_data', '[]')

            try:
                table_data = json.loads(table_data)
            except json.JSONDecodeError:
                return JsonResponse({'error': 'Invalid table data format'}, status=400)

            for entry in table_data:
                employee_code = entry.get('employee_code')
                amount = entry.get('amount')

                if not employee_code or not amount:
                    return JsonResponse({'error': 'Missing employee_code or amount'}, status=400)

                try:
                    employee = Employee.objects.get(employee_code=employee_code, status='active')
                except Employee.DoesNotExist:
                    return JsonResponse({'error': f'Employee with code {employee_code} does not exist or is not active'}, status=404)

                Incentive.objects.create(
                    employee=employee,
                    date=date,
                    incentive_amount=amount
                )

            return JsonResponse({'status': 'success'})
        else:
            return JsonResponse({'error': 'Form is not valid', 'details': form.errors.as_json()}, status=400)

    elif request.method == 'GET':
        employee_code = request.GET.get('employee_code')
        date = request.GET.get('date')

        if employee_code and date:
            employee = Employee.objects.filter(employee_code=employee_code, status='active').first()
            if employee:
                return JsonResponse({
                    'employee': {
                        'employee_code': employee.employee_code,
                        'employee_name': employee.employee_name
                    }
                })
            return JsonResponse({'error': 'Active employee not found'}, status=404)

        # For GET request, return an empty form and any error messages
        form = IncentiveForm()
        return render(request, 'employee/incentive/incentive_form.html', {'form': form})

    # Handle any other method (e.g., GET without required params)
    form = IncentiveForm()
    return render(request, 'employee/incentive/incentive_form.html', {'form': form})




def incentive_list_view(request):
    incentives = Incentive.objects.select_related('employee').all()
    return render(request, 'employee/incentive/incentive_list.html', {'incentives': incentives})

def incentive_update_view(request, incentive_id):
    incentive = get_object_or_404(Incentive, id=incentive_id)
    if request.method == 'POST':
        form = IncentiveForm(request.POST, instance=incentive)
        if form.is_valid():
            form.save()
            return redirect('incentive_list')  # Adjust the redirect as needed
    else:
        form = IncentiveForm(instance=incentive)
    return render(request, 'employee/incentive/incentive_form.html', {'form': form})

# views.py
from django.shortcuts import render, get_object_or_404, redirect
from .models import Incentive

def incentive_delete_view(request, incentive_id):
    incentive = get_object_or_404(Incentive, id=incentive_id)
    if request.method == 'POST':
        incentive.delete()
        return redirect('incentive_list')
    return render(request, 'employee/incentive/incentive_confirm_delete.html', {'incentive': incentive})

# views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from .models import Deduction, Employee
import json
from django.shortcuts import render, redirect
from django.http import JsonResponse
from .forms import DeductionForm, DeductionEditForm , DeductionAmountForm

def deduction_list_view(request):
    deductions = Deduction.objects.select_related('employee').all()
    return render(request, 'employee/deduction/deduction_list.html', {'deductions': deductions})

def deduction_create_view(request):
    if request.method == 'POST':
        deduction_form = DeductionForm(request.POST)
        
        # Debugging information to help understand form validity
        print('Form data:', request.POST)
        print('Form errors:', deduction_form.errors)
        
        if deduction_form.is_valid():
            date = deduction_form.cleaned_data['date']
            deduction_type = deduction_form.cleaned_data['deduction_type']
            table_data = request.POST.get('table_data', '[]')

            try:
                table_data = json.loads(table_data)
            except json.JSONDecodeError:
                return JsonResponse({'error': 'Invalid table data format'}, status=400)

            for entry in table_data:
                employee_code = entry.get('employee_code')
                amount = entry.get('amount')

                if not employee_code or not amount:
                    return JsonResponse({'error': 'Missing employee_code or amount'}, status=400)

                try:
                    employee = Employee.objects.get(employee_code=employee_code, status='active')
                except Employee.DoesNotExist:
                    return JsonResponse({'error': f'Employee with code {employee_code} does not exist or is not active'}, status=404)

                Deduction.objects.create(
                    employee=employee,
                    date=date,
                    deduction_type=deduction_type,
                    amount=amount
                )

            return JsonResponse({'status': 'success'})
        else:
            return JsonResponse({'error': 'Form is not valid', 'details': deduction_form.errors}, status=400)

    elif request.method == 'GET':
        employee_code = request.GET.get('employee_code')
        if employee_code:
            employee = Employee.objects.filter(employee_code=employee_code, status='active').first()
            if employee:
                return JsonResponse({
                    'employee': {
                        'employee_code': employee.employee_code,
                        'employee_name': employee.employee_name
                    }
                })
            return JsonResponse({'error': 'Active employee not found'}, status=404)
        else:
            deduction_form = DeductionForm()
            return render(request, 'employee/deduction/deduction_form.html', {
                'deduction_form': deduction_form,
                'error_message': 'No employee code provided'
            })

    return render(request, 'employee/deduction/deduction_form.html', {
        'deduction_form': DeductionForm()
    })

from django.shortcuts import get_object_or_404, redirect, render
from .models import Deduction
from .forms import DeductionEditForm

def deduction_update_view(request, deduction_id):
    # Retrieve the Deduction object or return 404 if not found
    deduction = get_object_or_404(Deduction, id=deduction_id)

    if request.method == 'POST':
        # Initialize the form with the POST data and the existing Deduction instance
        form = DeductionEditForm(request.POST, instance=deduction)
        
        if form.is_valid():
            # Save the form if it's valid
            form.save()
            # Redirect to the deduction list view after saving
            return redirect('deduction_list')
    else:
        # Initialize the form with the existing Deduction instance for GET request
        form = DeductionEditForm(instance=deduction)

    # Render the template with the form to display
    return render(request, 'employee/deduction/deduction_edit_form.html', {'form': form})


def deduction_delete_view(request, deduction_id):
    deduction = get_object_or_404(Deduction, id=deduction_id)
    if request.method == 'POST':
        deduction.delete()
        return redirect('deduction_list')
    return render(request, 'employee/deduction/deduction_confirm_delete.html', {'deduction': deduction})




def salary_reports(request):
    return render(request, 'employee/employee_salary/salary_reports.html')

    
def salary_abstract_reports(request):
    return render(request, 'employee/employee_salary/salary_abstract_reports.html')

    

from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Sum
from datetime import timedelta
from attendance.models import Attendance
from .forms import NonPfSalaryForm
from .utils import render_to_pdf  # Assuming you have a function to render PDFs
from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Sum
from datetime import datetime
from .forms import NonPfSalaryForm
from .utils import render_to_pdf  # Make sure this utility function exists
from django.shortcuts import render
from django.db.models import Sum
from datetime import datetime, timedelta
from .forms import NonPfSalaryForm
from decimal import Decimal
from django.shortcuts import render
from django.http import HttpResponse
from .forms import NonPfSalaryForm
from datetime import datetime, timedelta
from decimal import Decimal
from django.db.models import Sum
from io import BytesIO
import pandas as pd
from xhtml2pdf import pisa
from django.template.loader import get_template
from decimal import Decimal
from django.shortcuts import render
from django.db.models import Sum
from django.db.models.functions import Coalesce
from django.shortcuts import render, redirect
from datetime import datetime, timedelta
from decimal import Decimal
from .forms import NonPfSalaryForm


def non_pf_salary_view(request):
    form = NonPfSalaryForm(request.POST or None)
    salary_details = []

    # Initialize totals for all columns and deductions
    total_paid_days = 0
    total_ot_hours = Decimal(0)
    all_total_wages = Decimal(0)
    total_ot_amount = Decimal(0)
    total_gross_amount = Decimal(0)
    all_total_deductions=Decimal(0)
    total_deductions = Decimal(0)
    total_incentives = Decimal(0)
    total_net_salary = Decimal(0)
    total_bus_amount = Decimal(0)
    total_employee_food_expense = Decimal(0)
    total_advance_deductions = Decimal(0)
    total_canteen_deductions = Decimal(0)
    total_mess_deductions = Decimal(0)
    total_store_deductions = Decimal(0)
    total_other_deductions = Decimal(0)
    total_rent = Decimal(0)

    if form.is_valid():
        selected_date = form.cleaned_data['month']
        year = selected_date.year
        month = selected_date.month

        first_day_of_month = datetime(year, month, 1)
        last_day_of_month = (first_day_of_month.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

        attendance_records = Attendance.objects.filter(
            date__range=[first_day_of_month, last_day_of_month],
            employee__pf_esi_applicable='no',
            employee__status='active'
        )

        # Filter by category and department if provided
        if form.cleaned_data.get('category'):
            attendance_records = attendance_records.filter(employee__category=form.cleaned_data['category'])
        if form.cleaned_data.get('department'):
            attendance_records = attendance_records.filter(employee__department=form.cleaned_data['department'])

        # Group attendance records by employee
        employees = {}
        for record in attendance_records:
            if record.employee not in employees:
                employees[record.employee] = {
                    'attendance_records': [],
                    'total_p_days': 0,
                    'total_ot_hours': Decimal(0),
                    'incentives': Decimal(0),
                   
                    'deductions': {
                        'advance': Decimal(0),
                        'canteen': Decimal(0),
                        'mess': Decimal(0),
                        'store': Decimal(0),
                        'others': Decimal(0),
                        'total': Decimal(0),
                    }
                }

            if record.absent == 'P':
                employees[record.employee]['total_p_days'] += 1

            if record.ot_hours:
                employees[record.employee]['total_ot_hours'] += Decimal(record.ot_hours)

            employees[record.employee]['attendance_records'].append(record)

        # Calculate salary details for each employee
        for i, (employee, data) in enumerate(employees.items(), start=1):
            try:
                wages = Wages.objects.filter(
                    employee=employee,
                    start_date__lte=first_day_of_month,
                    end_date__gte=last_day_of_month
                ).first()
                per_day_salary = Decimal(wages.per_day) if wages else Decimal(0)
                hra_amount = Decimal(wages.hra_amount) if wages else Decimal(0)
                other_allowance = Decimal(wages.other_allowance) if wages else Decimal(0)
                rent = Decimal(wages.rent) if wages else Decimal(0)
            except Wages.DoesNotExist:
                per_day_salary = hra_amount = other_allowance = rent = Decimal(0)
            

            # Fetch and calculate incentives
            employee_incentives = Incentive.objects.filter(
                employee=employee,
                date__month=month,
                date__year=year
            ).aggregate(total_incentives=Coalesce(Sum('incentive_amount'), Decimal(0)))

            data['incentives'] = employee_incentives['total_incentives'] or Decimal(0)

            category_hours = Decimal(employee.category.hours) if employee.category else Decimal(8)
            ot_amount = int(round((per_day_salary / category_hours) * data['total_ot_hours'] * 1))
            total_wages = data['total_p_days'] * per_day_salary
            gross_amount = total_wages + ot_amount + rent + data['incentives'] 

                        # Calculate deductions
            deductions = Deduction.objects.filter(employee=employee, date__month=month, date__year=year)
            for deduction in deductions:
                if deduction.deduction_type in data['deductions']:
                    data['deductions'][deduction.deduction_type] += deduction.amount
                    if deduction.deduction_type == 'advance':
                        total_advance_deductions += deduction.amount
                    elif deduction.deduction_type == 'canteen':
                        total_canteen_deductions += deduction.amount
                    elif deduction.deduction_type == 'mess':
                        total_mess_deductions += deduction.amount
                    elif deduction.deduction_type == 'store':
                        total_store_deductions += deduction.amount
                    elif deduction.deduction_type == 'others':
                        total_other_deductions += deduction.amount

            total_deductions = sum(data['deductions'].values())
            
  
            net_salary = gross_amount - total_deductions 
            net_salary = round(net_salary)

            salary_details.append({
                'sl_no': i,
                'employee_code': employee.employee_code,
                'name': employee.employee_name,
                'department': employee.department.name,
                'designation': employee.designation.name if employee.designation else '',
                'hra_amount': float(hra_amount),
                'other_allowance': float(other_allowance),
                'rent': float(rent),
                'total_p_days': float(data['total_p_days']),
                'total_ot_hours': float(data['total_ot_hours']),
                'per_day_salary': float(per_day_salary),
                'total_wages': float(total_wages),
                'ot_amount': float(ot_amount),
                'gross_amount': float(gross_amount),
                'total_deductions': float(total_deductions),
                'incentives': float(data['incentives']),
                'net_salary': float(net_salary),
                
                'deductions': {key: float(value) for key, value in data['deductions'].items()},
            })
         # Sort the salary details by employee code
            salary_details = sorted(salary_details, key=lambda x: x['employee_code'])  # Sorting employee codes in ascending order  
        

            # Update overall totals
            total_paid_days += data['total_p_days']
            total_ot_hours += data['total_ot_hours']
            all_total_wages += total_wages
            total_ot_amount += ot_amount
            total_rent += Decimal(rent)  # Convert float rent to Decimal
            total_gross_amount += gross_amount
            all_total_deductions += total_deductions
            total_incentives += data['incentives']
            total_net_salary += net_salary

        # Save data in session for export
        request.session.update({
            'salary_details': salary_details,
            'month': selected_date.strftime('%Y-%m'),
            'total_paid_days': float(total_paid_days),
            'total_ot_hours': float(total_ot_hours),
            'all_total_wages': float(all_total_wages),
            'total_ot_amount': float(total_ot_amount),
            'total_gross_amount': float(total_gross_amount),
            'all_total_deductions': float(all_total_deductions),
            'total_rent':float(total_rent),
         
            'total_incentives': float(total_incentives),
            'total_net_salary': float(total_net_salary),
            'total_advance_deductions': float(total_advance_deductions),
            'total_canteen_deductions': float(total_canteen_deductions),
            'total_mess_deductions': float(total_mess_deductions),
            'total_store_deductions': float(total_store_deductions),
            'total_other_deductions': float(total_other_deductions),
        })

        return redirect(reverse('non_pf_salary_details_view'))

    return render(request, 'employee/employee_salary/non_pf_salary.html', {'form': form, 'salary_details': salary_details})

def non_pf_salary_details_view(request):
    salary_details = request.session.get('salary_details', [])
    selected_month = request.session.get('month', '')

    # Initialize total deductions
    total_wise_deductions = {
        'advance': 0,
        'mess': 0,
        'store': 0,
        'others': 0,
    }

    # Accumulate total deductions
    for detail in salary_details:
        total_wise_deductions['advance'] += detail['deductions'].get('advance', 0)
        total_wise_deductions['mess'] += detail['deductions'].get('mess', 0)
        total_wise_deductions['store'] += detail['deductions'].get('store', 0)
        total_wise_deductions['others'] += detail['deductions'].get('others', 0)

    # Retrieve session totals
    total_paid_days = request.session.get('total_paid_days', 0)
    total_ot_hours = request.session.get('total_ot_hours', 0)
    all_total_wages = request.session.get('all_total_wages', 0)
    total_ot_amount = request.session.get('total_ot_amount', 0)
    total_rent = request.session.get('total_rent', 0),
    total_gross_amount = request.session.get('total_gross_amount', 0)
    all_total_deductions = request.session.get('all_total_deductions', 0)
    total_incentives = request.session.get('total_incentives', 0)
    total_net_salary = request.session.get('total_net_salary', 0)


    # Render the details view
    return render(request, 'employee/employee_salary/non_pf_salary_details_view.html', {
        'salary_details': salary_details,
        'selected_month': selected_month,
        'total_paid_days': total_paid_days,
        'total_ot_hours': total_ot_hours,
        'all_total_wages': all_total_wages,
        'total_ot_amount': total_ot_amount,
        'total_gross_amount': total_gross_amount,
        'all_total_deductions': all_total_deductions,
        'total_incentives': total_incentives,
        'total_net_salary': total_net_salary,
        'total_rent' : total_rent,
        
        'total_wise_deductions': total_wise_deductions,
    })

import pandas as pd
from io import BytesIO
from django.http import HttpResponse

def non_pf_salary_download_excel(request):
    salary_details = request.session.get('salary_details', [])
    month = request.session.get('month', '')

    if not salary_details:
        return HttpResponse("No salary details to export.", status=400)

    # Prepare data for Excel export
    columns = [
        'Serial No', 'Employee Code', 'Employee Name', 'Department', 'Total Paid Days',
        'Total OT Hours', 'Per Day Salary', 'Total Wages', 'OT Amount', 'Gross Amount',
         'Total Incentives', 'Total Deductions', 'Net Salary'
    ]

    # Prepare the data for the DataFrame
    data = [
        [
            entry['sl_no'],
            entry['employee_code'],
            entry['name'],
            entry['department'],
            entry['total_p_days'],
            entry['total_ot_hours'],
            entry['per_day_salary'],
            entry['total_wages'],
            entry['ot_amount'],
            entry['gross_amount'],
           
            entry['incentives'],
            entry['total_deductions'],
            entry['net_salary']
        ] for entry in salary_details
    ]

    # Create DataFrame from data
    df = pd.DataFrame(data, columns=columns)

    # Create a summary row for totals (only applicable to certain columns)
    total_row = {
        'Serial No': 'Total',
        'Employee Code': '',
        'Employee Name': '',
        'Department': '',
        'Total Paid Days': df['Total Paid Days'].sum(),
        'Total OT Hours': df['Total OT Hours'].sum(),
        'Per Day Salary': '',
        'Total Wages': df['Total Wages'].sum(),
        'OT Amount': df['OT Amount'].sum(),
        'Gross Amount': df['Gross Amount'].sum(),
        
        'Total Incentives': df['Total Incentives'].sum(),
        'Total Deductions': df['Total Deductions'].sum(),
        'Net Salary': df['Net Salary'].sum()
    }

    # Append the total row to the DataFrame
    total_df = pd.DataFrame([total_row])
    df = pd.concat([df, total_df], ignore_index=True)

    # Write to Excel file
    with BytesIO() as buffer:
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Non PF Salary')
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="non_pf_salary_{month}.xlsx"'
        return response

def render_to_pdf(template_src, context_dict):
    template = get_template(template_src)
    html = template.render(context_dict)
    result = BytesIO()
    pdf = pisa.CreatePDF(BytesIO(html.encode("UTF-8")), dest=result)
    if not pdf.err:
        return result.getvalue()
    return None


def non_pf_salary_download_pdf(request):
    # Retrieve salary details from the session
    salary_details = request.session.get('salary_details', [])
    month = request.session.get('month', '')

    # Check if salary details are available
    if not salary_details:
        return HttpResponse("No salary details to export.", status=400)

    # Prepare context for the PDF
    context = {
        'salary_details': salary_details,
        'month': month,
        'total_paid_days': request.session.get('total_paid_days', ),
        'total_ot_hours': request.session.get('total_ot_hours', ),
        'all_total_wages': request.session.get('all_total_wages', ),
        'total_ot_amount': request.session.get('total_ot_amount', ),
        'total_gross_amount': request.session.get('total_gross_amount', ),
        'total_deductions': request.session.get('total_deductions', ),
        'total_incentives': request.session.get('total_incentives', ),
        'total_net_salary': request.session.get('total_net_salary', ),
       

        # Individual deduction totals
        'all_total_deductions':request.session.get('all_total_deductions', ),
        'total_advance_deductions': request.session.get('total_advance_deductions', ),
        'total_canteen_deductions': request.session.get('total_canteen_deductions', ),
        'total_mess_deductions': request.session.get('total_mess_deductions', ),
        'total_store_deductions': request.session.get('total_store_deductions', ),
        'total_other_deductions': request.session.get('total_other_deductions', ),
        
        # Calculate total deductions
        'total_deductions_details': (
            request.session.get('total_advance_deductions', ) +
            request.session.get('total_canteen_deductions', ) +
            request.session.get('total_mess_deductions', ) +
            request.session.get('total_store_deductions', ) +
            request.session.get('total_other_deductions', )
        )
    }

    # Generate PDF
    pdf = render_to_pdf('employee/employee_salary/non_pf_salary_pdf.html', context)
    if pdf:
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="non_pf_salary_{month}.pdf"'
        return response

    return HttpResponse("Error generating PDF", status=400)


import pandas as pd
from io import BytesIO
from django.http import HttpResponse

def employee_pf_salary_for_cover(request):
    # Retrieve salary details from the session
    salary_details = request.session.get('salary_details', [])

    # Return an error if no salary details are found
    if not salary_details:
        return HttpResponse("No salary details to export.", status=400)

    # Define the base columns for the DataFrame with Employee Code for both sets
    columns = [
        'Sl No', 'Employee Code', 'Name', 'Net Salary',  # First set
    ]
    
    # Prepare the data in pairs of columns
    data = []
    total_net_salary = 0  # To keep track of the total net salary for the last row

    for i in range(0, len(salary_details)):
        salary = salary_details[i]
        total_net_salary += salary.get('net_salary', 0)  # Accumulate net salary

        row = [
            i + 1,  # 'Sl No' (Index starts from 1)
            salary.get('employee_code', ''),  # Employee Code
            salary.get('name', ''),  # Employee Name
            salary.get('net_salary', 0)  # Net Salary
        ]
        data.append(row)

    # Add a row for the totals
    data.append([
        "Total", "", "", total_net_salary  # Total row, no employee code or name
    ])

    # Create a DataFrame with the formatted data
    df = pd.DataFrame(data, columns=columns)

    # Write the DataFrame to an Excel file
    with BytesIO() as buffer:
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Employee Salary Cover')

        # Return the Excel file as a response
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="employee_salary_cover.xlsx"'
        return response

# Import statements
from django.shortcuts import render
from django.urls import reverse
from django.http import HttpResponse
from io import BytesIO
from datetime import datetime, timedelta
from decimal import Decimal
from django.db.models import Sum
from django.template.loader import render_to_string, get_template
from xhtml2pdf import pisa
import pandas as pd
from .forms import ExtraDaysSalaryForm
from decimal import Decimal, getcontext


# Import statements
from django.shortcuts import render
from django.urls import reverse
from django.http import HttpResponse
from io import BytesIO
from datetime import datetime, timedelta
from decimal import Decimal
from django.db.models import Sum
from django.template.loader import render_to_string, get_template
from xhtml2pdf import pisa
import pandas as pd
from .forms import ExtraDaysSalaryForm
from decimal import Decimal, getcontext
from django.db.models import Q
from decimal import Decimal
from datetime import datetime, timedelta
from django.shortcuts import render, redirect
from django.urls import reverse
from audit.models import Holiday

def extra_days_salary_view(request):
    form = ExtraDaysSalaryForm(request.POST or None)
    salary_details = []
    
    # Initialize total variables
    total_extra_days_wages = 0
    total_ot_amount = 0
    total_incentives = 0
    total_food_expense = 0
    total_bus_fare = 0
    total_net_salary = 0
    total_extra_days = 0
    total_ot_hours = 0
    total_per_day_salary = 0
    total_rent = 0  # Add total_rent to accumulate the rent values

    if form.is_valid():
        selected_date = form.cleaned_data['month']
        year = selected_date.year
        month = selected_date.month

        first_day_of_month = datetime(year, month, 1)
        last_day_of_month = (first_day_of_month.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

        employees = Employee.objects.filter(pf_esi_applicable='yes', status='active')

        # Filter employees based on form data
        if form.cleaned_data.get('category'):
            employees = employees.filter(category=form.cleaned_data['category'])
        if form.cleaned_data.get('department'):
            employees = employees.filter(department=form.cleaned_data['department'])

        # Get attendance records for all employees in the selected date range
        attendance_records = Attendance.objects.filter(
            date__range=[first_day_of_month, last_day_of_month],
            employee__in=employees
        )

        employees_dict = {}
        for employee in employees:
            employees_dict[employee] = {
                'extra_days': 0,
                'total_ot_hours': Decimal(0),
                'incentives': Decimal(0),
                'weekly_off_day': employee.weekly_off,
                'total_present_days': 0,
            }

        for record in attendance_records:
            employee = record.employee
            weekly_off_day = employees_dict[employee]['weekly_off_day']

            # Check if this weekly off day is also a national or festival holiday
            if record.date.strftime('%A') == weekly_off_day:
                is_holiday = Holiday.objects.filter(
                    Q(date=record.date),
                    Q(holiday_type='NH') | Q(holiday_type='FH')  # Only consider National and Festival Holidays
                ).exists()

                # If it's not a holiday and the employee was present, count it as an extra day
                if not is_holiday and record.absent == 'P':
                    employees_dict[employee]['extra_days'] += 1

            # If the employee was present, count it as a present day
            if record.absent == 'P':
                employees_dict[employee]['total_present_days'] += 1

            # Add overtime hours if applicable
            if record.ot_hours:
                ot_hours = Decimal(record.ot_hours)
                employees_dict[employee]['total_ot_hours'] += ot_hours

        # Calculate wages, incentives, and other details for each employee
        for i, (employee, data) in enumerate(employees_dict.items(), start=1):
            try:
                # Fetch the wages for the employee
                wages = Wages.objects.get(employee=employee, start_date__lte=first_day_of_month, end_date__gte=last_day_of_month)
                per_day_salary = Decimal(wages.per_day)
                rent = Decimal(wages.rent)
            except Wages.DoesNotExist:
                per_day_salary = Decimal(0)
                rent = Decimal(0)

            # Calculate extra days wages
            extra_days_wages = data['extra_days'] * per_day_salary

            # Get the number of hours in the employee's category
            category_hours = Decimal(employee.category.hours) if employee.category else Decimal(0)
            ot_amount = int(round((per_day_salary / category_hours) * data['total_ot_hours'] * 1))
            employee_incentives = Incentive.objects.filter(
                employee=employee,
                date__month=month,
                date__year=year
            ).aggregate(total_incentives=Coalesce(Sum('incentive_amount'), Decimal(0)))

            data['incentives'] = employee_incentives['total_incentives'] or Decimal(0)

            # Calculate net salary
            net_salary = extra_days_wages + ot_amount + data['incentives'] + rent
            net_salary = round(net_salary)  # Round net salary to the nearest 10

            # Update total variables for the session
            total_extra_days_wages += float(extra_days_wages)
            total_ot_amount += float(ot_amount)
            total_incentives += float(data['incentives'])
            total_net_salary += float(net_salary)
            total_rent += float(rent)  # Accumulate rent into total_rent

            # Accumulate extra days, total OT hours, and per day salary
            total_extra_days += data['extra_days']
            total_ot_hours += float(data['total_ot_hours'])
            total_per_day_salary += float(per_day_salary)

            # Append salary details for each employee
            salary_details.append({
                'sl_no': i,
                'employee_code': employee.employee_code,
                'employee_name': employee.employee_name,
                'name': employee.employee_name,
                'department': employee.department.name if employee.department else '',
                'designation': employee.designation.name if employee.designation else '',
                'extra_days': data['extra_days'],
                'total_ot_hours': float(data['total_ot_hours']),
                'per_day_salary': float(per_day_salary),
                'rent': float(rent),  # Include rent in details
                'extra_days_wages': float(extra_days_wages),
                'ot_amount': float(ot_amount),
                'total_incentives': float(data['incentives']),
                'net_salary': float(net_salary),
            })

        # Sort the salary details by employee code
        salary_details = sorted(salary_details, key=lambda x: x['employee_code'])

        # Save the salary details and totals in the session
        request.session['salary_details'] = salary_details
        request.session['total_extra_days_wages'] = total_extra_days_wages
        request.session['total_ot_amount'] = total_ot_amount
        request.session['total_incentives'] = total_incentives
        request.session['total_net_salary'] = total_net_salary
        request.session['total_rent'] = total_rent  # Save total_rent in session
        request.session['total_extra_days'] = total_extra_days
        request.session['total_ot_hours'] = total_ot_hours
        request.session['total_per_day_salary'] = total_per_day_salary
        request.session['month'] = selected_date.strftime('%Y-%m')

        return redirect(reverse('extra_salary_details_view'))

    return render(request, 'employee/employee_salary/extra_days_salary.html', {
        'form': form,
        'salary_details': salary_details,
    })



def extra_salary_details_view(request):
    salary_details = request.session.get('salary_details', [])
    selected_month = request.session.get('month', '')

    # Retrieve the totals from the session
    total_extra_days_wages = request.session.get('total_extra_days_wages', 0)
    total_ot_amount = request.session.get('total_ot_amount', 0)
    total_incentives = request.session.get('total_incentives', 0)
    total_net_salary = request.session.get('total_net_salary', 0)
    total_extra_days = request.session.get('total_extra_days', 0)
    total_ot_hours = request.session.get('total_ot_hours', 0)
    total_per_day_salary = request.session.get('total_per_day_salary', 0)
    total_rent = request.session.get('total_rent', 0)  # Retrieve total_rent

    return render(request, 'employee/employee_salary/extra_salary_details_view.html', {
        'salary_details': salary_details,
        'selected_month': selected_month,
        'total_extra_days_wages': total_extra_days_wages,
        'total_ot_amount': total_ot_amount,
        'total_incentives': total_incentives,
        'total_net_salary': total_net_salary,
        'total_extra_days': total_extra_days,
        'total_ot_hours': total_ot_hours,
        'total_per_day_salary': total_per_day_salary,
        'total_rent': total_rent,  # Include total_rent in the template
    })
def render_to_pdf(template_src, context_dict):
    template = get_template(template_src)
    html = template.render(context_dict)
    result = BytesIO()
    pdf = pisa.CreatePDF(BytesIO(html.encode("UTF-8")), dest=result)
    if not pdf.err:
        return result.getvalue()
    return None

def extra_salary_download_pdf(request):
    salary_details = request.session.get('salary_details', [])
    month = request.session.get('month', '')

    if not salary_details:
        return HttpResponse("No salary details to export.", status=400)

    # Retrieve the totals from the session
    total_extra_days_wages = request.session.get('total_extra_days_wages', 0)
    total_ot_amount = request.session.get('total_ot_amount', 0)
    total_incentives = request.session.get('total_incentives', 0)

    total_net_salary = request.session.get('total_net_salary', 0)
    total_extra_days = request.session.get('total_extra_days', 0)
    total_ot_hours = request.session.get('total_ot_hours', 0)
    total_per_day_salary = request.session.get('total_per_day_salary', 0)
    total_rent = request.session.get('total_rent', 0)  # Retrieve total_rent

    context = {
        'salary_details': salary_details,
        'month': month,
        'total_extra_days_wages': total_extra_days_wages,
        'total_ot_amount': total_ot_amount,
        'total_incentives': total_incentives,
       
        'total_net_salary': total_net_salary,
        'total_extra_days': total_extra_days,
        'total_ot_hours': total_ot_hours,
        'total_per_day_salary': total_per_day_salary,
        'total_rent': total_rent,  # Include total_rent in the template
    }

    pdf = render_to_pdf('employee/employee_salary/extra_days_salary_pdf.html', context)
    if pdf:
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="extra_days_salary_{month}.pdf"'
        return response
    return HttpResponse("Error generating PDF", status=400)



import pandas as pd
from io import BytesIO
from django.http import HttpResponse

def extra_salary_download_excel(request):
    salary_details = request.session.get('salary_details', [])
    month = request.session.get('month', '')

    if not salary_details:
        return HttpResponse("No salary details to export.", status=400)

    # Prepare data for Excel export
    columns = ['Serial No', 'Employee Code', 'Employee Name', 'Department',
               'Extra Days', 'Total OT Hours', 'Per Day Salary', 'Extra Days Wages',
               'OT Amount', 'Total Incentives', 'Net Salary']

    data = [
        [
            entry['sl_no'],
            entry['employee_code'],
            entry['employee_name'],
            entry['department'],
            entry['extra_days'],
            entry['total_ot_hours'],
            entry['per_day_salary'],
            entry['extra_days_wages'],
            entry['ot_amount'],
            entry['total_incentives'],
            entry['net_salary']
        ] for entry in salary_details
    ]

    df = pd.DataFrame(data, columns=columns)

    # Create a summary row for totals
    total_row = {
        'Serial No': 'Total',
        'Employee Code': '',
        'Employee Name': '',
        'Department': '',
        'Extra Days': df['Extra Days'].sum(),
        'Total OT Hours': df['Total OT Hours'].sum(),
        'Per Day Salary': '',
        'Extra Days Wages': df['Extra Days Wages'].sum(),
        'OT Amount': df['OT Amount'].sum(),
        'Total Incentives': df['Total Incentives'].sum(),
        'Net Salary': df['Net Salary'].sum(),
    }

    # Create a DataFrame for the total row
    total_df = pd.DataFrame([total_row])

    # Concatenate the original DataFrame with the total DataFrame
    df = pd.concat([df, total_df], ignore_index=True)

    # Write to an Excel file
    with BytesIO() as buffer:
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Extra Days Salary')
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="extra_days_salary_{month}.xlsx"'
        return response




###############################################################################################################################
# views.py

from django.shortcuts import render
from .forms import DeductionFilterForm
from .models import Deduction

def deduction_filterd_list(request):
    form = DeductionFilterForm(request.POST or None)
    deductions = Deduction.objects.none()

    if form.is_valid():
        start_date = form.cleaned_data['start_date']
        end_date = form.cleaned_data['end_date']
        deduction_type = form.cleaned_data['deduction_type']

        # Filter deductions based on the form inputs
        deductions = Deduction.objects.filter(date__range=[start_date, end_date])
        
        if deduction_type:
            deductions = deductions.filter(deduction_type=deduction_type)

    return render(request, 'employee/deduction/deduction_filterd_list.html', {
        'form': form,
        'deductions': deductions,
    })



# views.py

import openpyxl
from django.shortcuts import render, redirect
from django.contrib import messages
from .forms import EmployeePFESIUploadForm
from .models import Employee

def employee_pfesi_upload_view(request):
    if request.method == 'POST':
        form = EmployeePFESIUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = form.cleaned_data['excel_file']

            # Open the Excel file
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            # Iterate through the rows in the Excel file
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
                employee_code, pf_esi_applicable, pf_no, esi_no, pf_date_of_joining = row

                try:
                    employee = Employee.objects.get(employee_code=employee_code)
                    employee.pf_esi_applicable = pf_esi_applicable
                    employee.pf_no = pf_no
                    employee.esi_no = esi_no
                    employee.pf_date_of_joining = pf_date_of_joining

                    # Save the updated employee
                    employee.save()
                except Employee.DoesNotExist:
                    messages.warning(request, f"Employee with code {employee_code} does not exist.")
                    continue

            messages.success(request, "Employees' PF/ESI details have been updated successfully.")
            return redirect('employee_pfesi_upload_view')
    else:
        form = EmployeePFESIUploadForm()

    return render(request, 'employee/employee_pfesi_upload.html', {'form': form})



import pandas as pd
from django.shortcuts import render, redirect
from django.contrib import messages
from .forms import EmployeeUploadForm
from .models import Employee
import pandas as pd
from django.shortcuts import render, redirect
from django.contrib import messages
from .forms import EmployeeDatesUploadForm # Ensure EmployeeUploadForm is correctly defined
from .models import Employee

def upload_employee_dates(request):
    if request.method == 'POST':
        form = EmployeeDatesUploadForm(request.POST, request.FILES)
        
        if form.is_valid():
            try:
                # Access the uploaded file from request.FILES
                if 'excel_file' in request.FILES:
                    excel_file = request.FILES['excel_file']
                    
                    # Load the Excel file into a pandas DataFrame using openpyxl
                    df = pd.read_excel(excel_file, engine='openpyxl')
                    
                    # Iterate through each row in the Excel file
                    for _, row in df.iterrows():
                        employee_code = row.get('Employee Code')
                        
                        # Skip if employee_code is not available
                        if not employee_code:
                            messages.warning(request, "Missing Employee Code in a row. Skipping this row.")
                            continue
                        
                        # Check if the employee exists in the database
                        try:
                            employee = Employee.objects.get(employee_code=employee_code)
                            
                            # Update only the date fields from the Excel file
                            employee.date_of_birth = row.get('Date of Birth') if pd.notnull(row.get('Date of Birth')) else None
                            employee.date_of_joining = row.get('Date of Joining') if pd.notnull(row.get('Date of Joining')) else None
                            employee.date_of_re_joining = row.get('Date of Re-joining') if pd.notnull(row.get('Date of Re-joining')) else None
                            employee.date_of_leaving = row.get('Date of Leaving') if pd.notnull(row.get('Date of Leaving')) else None
                            employee.pf_date_of_joining = row.get('PF Date of Joining') if pd.notnull(row.get('PF Date of Joining')) else None
                            
                            # Save the updated employee record
                            employee.save()
                        
                        except Employee.DoesNotExist:
                            messages.error(request, f"Employee with code {employee_code} does not exist.")
                    
                    messages.success(request, "Employee dates updated successfully.")
                    return redirect('employee_list')  # Redirect to the desired page after upload
                else:
                    messages.error(request, "No file uploaded.")
            except Exception as e:
                messages.error(request, f"Error processing file: {str(e)}")
        else:
            messages.error(request, "Invalid form submission.")
    else:
        form = EmployeeDatesUploadForm()

    return render(request, 'employee/upload_employee_dates.html', {'form': form})



import pandas as pd


def bulk_upload_deductions(request):
    if request.method == 'POST':
        # Get the uploaded Excel file from the form
        deductions_file = request.FILES.get('deductions_file')

        if deductions_file:
            try:
                # Read the Excel file into a Pandas DataFrame
                df = pd.read_excel(deductions_file)

                # Track whether any valid records were processed
                records_processed = False

                # Iterate through each row in the DataFrame
                for _, row in df.iterrows():
                    employee_code = row.get('Employee Code')
                    date = row.get('Date')
                    deduction_type = row.get('Deduction Type')
                    amount = row.get('Amount')

                    # Skip the row if employee_code is missing
                    if not employee_code:
                        messages.warning(request, "Employee Code is missing in a row. This row will be skipped.")
                        continue

                    # Check if the employee exists in the database
                    try:
                        employee = Employee.objects.get(employee_code=employee_code)
                    except Employee.DoesNotExist:
                        messages.warning(request, f"Employee code {employee_code} not found. Record skipped.")
                        continue

                    # Validate deduction type
                    if deduction_type not in dict(Deduction.DEDUCTION_TYPE_CHOICES).keys():
                        messages.warning(request, f"Invalid deduction type {deduction_type}. Record skipped.")
                        continue

                    # Create or update the deduction record
                    Deduction.objects.update_or_create(
                        employee=employee,
                        date=date,
                        deduction_type=deduction_type,
                        defaults={'amount': amount}
                    )

                    records_processed = True

                # Display appropriate messages based on the records processed
                if records_processed:
                    messages.success(request, "Deduction records processed successfully.")
                else:
                    messages.info(request, "No valid records were uploaded.")

            except Exception as e:
                messages.error(request, f"Error processing the file: {e}")

            # Redirect to the bulk upload page after processing
            return redirect('bulk_upload_deductions')

    # Render the bulk upload form if the request is not a POST
    return render(request, 'employee/deduction/bulk_upload_deductions.html')

###############################################################################################################################

###############################################################################################################################




from decimal import Decimal
from datetime import datetime, timedelta
from django.shortcuts import render
from .forms import NonPfCategorySalaryForm
from decimal import Decimal
from django.db.models import Sum
from datetime import datetime, timedelta
from datetime import datetime, timedelta
from decimal import Decimal
from django.shortcuts import render
from .forms import NonPfCategorySalaryForm
from datetime import datetime

from decimal import Decimal

# Helper function to convert Decimal to float
def convert_decimals_to_float(data):
    if isinstance(data, dict):
        return {key: convert_decimals_to_float(value) for key, value in data.items()}
    elif isinstance(data, list):
        return [convert_decimals_to_float(item) for item in data]
    elif isinstance(data, Decimal):
        return float(data)
    return data


def category_totals_view(request):
    form = NonPfCategorySalaryForm(request.POST or None)
    salary_details = []
    selected_month_name = None
    
    # Initialize overall totals
    total_paid_days = 0
    total_ot_hours = Decimal(0)
    all_total_wages = Decimal(0)
    total_ot_amount = Decimal(0)
    total_gross_amount = Decimal(0)
    all_total_deductions = Decimal(0)
    total_incentives = Decimal(0)
    total_advance_deductions = Decimal(0)
    total_canteen_deductions = Decimal(0)
    total_mess_deductions = Decimal(0)
    total_store_deductions = Decimal(0)
    total_other_deductions = Decimal(0)
    total_net_salary = Decimal(0)
    total_ot_amount = Decimal(0)
     

    # Initialize category and department totals
    category_totals = {}
    department_totals = {}

    # Overall totals for the organization
    overall_totals = {
        'total_employees': Decimal(0),
        'total_advance_deductions': Decimal(0),
        'total_mess_deductions': Decimal(0),
        'total_store_deductions': Decimal(0),
        'total_other_deductions': Decimal(0),
        'total_deductions': Decimal(0),
        'net_salary_total': Decimal(0),
        'total_incentives': Decimal(0),
        'total_gross_amount': Decimal(0),
        'total_days':Decimal(0),
        'all_employees':Decimal(0),
        'total_per_day_salary':Decimal(0),
        'total_shift_wages':Decimal(0),
        'total_ot_amount':Decimal(0),
        'total_rent':Decimal(0),
        'total_ot_hours':Decimal(0)

    }

    # Initialize the total employees count for the entire organization
    overall_total_employees = 0
    selected_month = None
    selected_year = None
    month = None
    year = None
    month_name = None

    if form.is_valid():
        selected_date = form.cleaned_data['month']
        year = selected_date.year
        month = selected_date.month
        selected_year = selected_date.year
        selected_month = selected_date.month
        month_name = selected_date.strftime('%B')
  

        # Define the date range for attendance records
        first_day_of_month = datetime(year, month, 1)
        last_day_of_month = (first_day_of_month.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

        # Query attendance records based on filters
        attendance_records = Attendance.objects.filter(
            date__range=[first_day_of_month, last_day_of_month],
            employee__pf_esi_applicable='no',
            employee__status='active'
        )

        # Apply category and department filters if provided
        if form.cleaned_data.get('category'):
            attendance_records = attendance_records.filter(employee__category=form.cleaned_data['category'])
        if form.cleaned_data.get('department'):
            attendance_records = attendance_records.filter(employee__department=form.cleaned_data['department'])

        # Group attendance records by employee
        employees = {}
        for record in attendance_records:
            if record.employee not in employees:
                employees[record.employee] = {
                    'attendance_records': [],
                    'total_p_days': 0,
                    'total_ot_hours': Decimal(0),
                    'incentives': Decimal(0),
                    'bus_fare': Decimal(0),
                    'food_expense': Decimal(0),
                    'deductions': {
                        'advance': Decimal(0),
                        'canteen': Decimal(0),
                        'mess': Decimal(0),
                        'store': Decimal(0),
                        'others': Decimal(0),
                        'total': Decimal(0),
                    }
                }

            # Count present days and overtime hours
            if record.absent == 'P':
                employees[record.employee]['total_p_days'] += 1
            if record.ot_hours:
                employees[record.employee]['total_ot_hours'] += Decimal(record.ot_hours)
            employees[record.employee]['attendance_records'].append(record)

        # Calculate salary details for each employee
        for i, (employee, data) in enumerate(employees.items(), start=1):
            try:
                wages = Wages.objects.filter(
                    employee=employee,
                    start_date__lte=first_day_of_month,
                    end_date__gte=last_day_of_month
                ).first()
                
                # Get wage details or set to zero if not found
                per_day_salary = Decimal(wages.per_day) if wages else Decimal(0)
                hra_amount = Decimal(wages.hra_amount) if wages else Decimal(0)
                other_allowance = Decimal(wages.other_allowance) if wages else Decimal(0)
                rent = Decimal(wages.rent) if wages else Decimal(0)
            except Wages.DoesNotExist:
                per_day_salary = hra_amount = other_allowance = rent = Decimal(0)
            overall_total_employees += 1
            # Fetch and calculate incentives
            employee_incentives = Incentive.objects.filter(
                employee=employee,
                date__month=month,
                date__year=year
            ).aggregate(total_incentives=Coalesce(Sum('incentive_amount'), Decimal(0)))

            data['incentives'] = employee_incentives['total_incentives'] or Decimal(0)
            # Calculate wages and gross amount
            category_hours = Decimal(employee.category.hours) if employee.category else Decimal(8)
            
            ot_amount = (per_day_salary / category_hours) * data['total_ot_hours']
            ot_amount = round(ot_amount)
            total_wages = data['total_p_days'] * per_day_salary
            gross_amount = total_wages + ot_amount + rent  + data['incentives'] 
            
           
            # Calculate deductions
            deductions = Deduction.objects.filter(employee=employee, date__month=month, date__year=year)
            for deduction in deductions:
                if deduction.deduction_type in data['deductions']:
                    data['deductions'][deduction.deduction_type] += deduction.amount
                    
                    # Track total deductions by type
                    if deduction.deduction_type == 'advance':
                        total_advance_deductions += deduction.amount
                    elif deduction.deduction_type == 'mess':
                        total_mess_deductions += deduction.amount
                    elif deduction.deduction_type == 'store':
                        total_store_deductions += deduction.amount
                    elif deduction.deduction_type == 'others':
                        total_other_deductions += deduction.amount

            # Calculate total deductions and net salary
            total_deductions = sum(data['deductions'].values())
            net_salary = gross_amount - total_deductions
            net_salary = round(net_salary)  # Round to nearest 10

            # Store salary details in a list
            salary_details.append({
                'sl_no': i,
                'employee_code': employee.employee_code,
                'name': employee.employee_name,
                'department': employee.department.name,
                'designation': employee.designation.name if employee.designation else '',
                'hra_amount': float(hra_amount),
                'other_allowance': float(other_allowance),
                'rent': float(rent),
                'total_p_days': float(data['total_p_days']),
                'total_ot_hours': float(data['total_ot_hours']),
                'per_day_salary': float(per_day_salary),
                'total_wages': float(total_wages),
                'ot_amount': float(ot_amount),
                'gross_amount': float(gross_amount),
                'total_deductions': float(total_deductions),
                'incentives': float(data['incentives']),
                'net_salary': float(net_salary),
                'bus_fare': float(data['bus_fare']),
                'food_expense': float(data['food_expense']),
                'deductions': {key: float(value) for key, value in data['deductions'].items()},
            })

            # Update overall totals
            total_paid_days += data['total_p_days']
            total_ot_hours += data['total_ot_hours']
            all_total_wages += total_wages
            total_ot_amount += ot_amount
            total_gross_amount += gross_amount
            all_total_deductions += total_deductions
            total_incentives += data['incentives']
            total_net_salary += net_salary  # Accumulate net salary here
            total_ot_amount += ot_amount
            
        
            # Accumulate department totals (new logic)
            department_name = employee.department.name if employee.department else 'Uncategorized'
            if department_name not in department_totals:
                department_totals[department_name] = {
                    'total_per_day_salary': Decimal(0),  # Track total per day salary
                    'total_employees': 0,  # Track the total number of employees
                    'shift_wages':0, 
                    'ot_amount':0,
                    'advance_total': Decimal(0),  # Track total advance deductions
                    'canteen_total': Decimal(0),  # Track total canteen deductions
                    'mess_total': Decimal(0),  # Track total mess deductions
                    'store_total': Decimal(0),  # Track total store deductions
                    'others_total': Decimal(0),  # Track total other deductions
                    'total_deductions': Decimal(0),  # Track total deductions in the department
                    'net_salary_total': Decimal(0),
                    'total_incentives': Decimal(0),
                    'total_gross_amount': Decimal(0),
                    'total_paid_days': 0,  # Ensure this key is initialized
                    'total_rent': 0, 
                    'total_ot_hours':0,
                }
            
            # Update department totals for this employee
            department_totals[department_name]['total_ot_hours'] += data['total_ot_hours']
            department_totals[department_name]['ot_amount'] += ot_amount
            department_totals[department_name]['shift_wages'] += total_wages
            department_totals[department_name]['total_per_day_salary'] += per_day_salary
            department_totals[department_name]['total_rent'] += rent
            department_totals[department_name]['total_paid_days'] += data['total_p_days']
            department_totals[department_name]['total_employees'] += 1
            department_totals[department_name]['advance_total'] += data['deductions']['advance']
            department_totals[department_name]['canteen_total'] += data['deductions']['canteen']
            department_totals[department_name]['mess_total'] += data['deductions']['mess']
            department_totals[department_name]['store_total'] += data['deductions']['store']
            department_totals[department_name]['others_total'] += data['deductions']['others']
            department_totals[department_name]['total_deductions'] += total_deductions
            department_totals[department_name]['net_salary_total'] += net_salary
            department_totals[department_name]['total_incentives'] += data['incentives'] 
            department_totals[department_name]['total_gross_amount'] += gross_amount
     

            # Update overall totals for the entire organization
             #:Decimal(0),
        #'all_employees':Decimal(0)
            overall_totals['total_ot_hours']  += data['total_ot_hours']
            overall_totals['total_rent']  +=  rent
            overall_totals['total_ot_amount']  +=  ot_amount
            overall_totals['total_shift_wages']  += total_wages
            overall_totals['total_days']  += data['total_p_days']
            overall_totals['total_per_day_salary'] += per_day_salary
            overall_totals['total_advance_deductions'] += data['deductions']['advance']
            overall_totals['total_mess_deductions'] += data['deductions']['mess']
            overall_totals['total_store_deductions'] += data['deductions']['store']
            overall_totals['total_other_deductions'] += data['deductions']['others']
            overall_totals['total_deductions'] += total_deductions
            overall_totals['net_salary_total'] += net_salary
            overall_totals['total_incentives'] += data['incentives'] 
            overall_totals['total_gross_amount'] += gross_amount

        # Update the overall employee count
            overall_totals['total_employees'] = overall_total_employees

    department_totals = dict(sorted(department_totals.items()))

    request.session['department_totals'] = convert_decimals_to_float(department_totals)
    request.session['overall_totals'] = convert_decimals_to_float(overall_totals)
    request.session['selected_month'] = month_name  # Store the selected month
    request.session['selected_year'] = year 
    # Render the response
    return render(request, 'employee/employee_salary/category_department_salary.html', {
        'form': form,
        'salary_details': salary_details,
        'category_totals': category_totals,
        'department_totals': department_totals,
        'overall_totals': overall_totals,
        'overall_total_employees': overall_total_employees,
        'selected_month': month_name,  # Pass selected month
    'selected_year': year,  # Pass selected year
        
    })




from django.shortcuts import redirect
from django.contrib import messages
from .models import NonPFAbstract

def save_non_pf_abstract(request):
    if request.method == "POST":
        # Retrieve data from the session
        department_totals = request.session.get('department_totals', {})
        selected_month = request.session.get('selected_month', 'Unknown')
        selected_year = request.session.get('selected_year', 'Unknown')

        # Ensure year is an integer
        try:
            selected_year = int(selected_year)
        except ValueError:
            messages.error(request, "Invalid year format.")
            return redirect('category_department_salary')  # Replace with your view name

        # Iterate over departments and save or update records
        for department_name, data in department_totals.items():
            obj, created = NonPFAbstract.objects.update_or_create(
                month=selected_month,
                year=selected_year,
                department_name=department_name,
                defaults={
                    'total_employees': data.get('total_employees', 0),
                    'total_paid_days': data.get('total_paid_days', 0),
                    'total_ot_hours': data.get('total_ot_hours', 0),
                    'ot_amount': data.get('ot_amount', 0.0),
                    'shift_wages': data.get('shift_wages', 0.0),
                    'per_day_salary': data.get('total_per_day_salary', 0.0),
                    'rent': data.get('total_rent', 0.0),
                    'advance_total': data.get('advance_total', 0.0),
                    'canteen_total': data.get('canteen_total', 0.0),
                    'mess_total': data.get('mess_total', 0.0),
                    'store_total': data.get('store_total', 0.0),
                    'others_total': data.get('others_total', 0.0),
                    'total_deductions': data.get('total_deductions', 0.0),
                    'total_gross_amount': data.get('total_gross_amount', 0.0),
                    'net_salary': data.get('net_salary_total', 0.0),
                    'total_incentives': data.get('total_incentives', 0.0),
                }
            )
            if created:
                messages.success(request, f"Created record for {department_name} ({selected_month} {selected_year}).")
            else:
                messages.info(request, f"Updated record for {department_name} ({selected_month} {selected_year}).")

        # Redirect to the appropriate page
        messages.success(request, "All department records have been successfully saved.")
        return redirect('category_department_salary_view')  # Replace with your view name
    else:
        # If accessed via GET, redirect to another page
        return redirect('category_department_salary_view')  # Replace with your view name


import openpyxl
from openpyxl.styles import Alignment, Font
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Alignment, Font
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from django.http import HttpResponse

def export_category_totals_excel(request):
    # Initialize workbook and worksheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Category Totals"
    
    # Set headers
    headers = [
        "Sl.No", "Department Name", "Total Employees", "Total Per Day Salary",
        "Total Paid Days", "Total OT Hours", "Shift Wages", "OT Amount",
        "Total Rent", "Total Incentives", "Total Gross Amount",
        "Advance Deductions", "Mess Deductions", "Store Deductions",
        "Other Deductions", "Net Salary"
    ]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Retrieve data from view logic
    department_totals = request.session.get('department_totals', {})
    overall_totals = request.session.get('overall_totals', {})

    # Write department data
    row_num = 2
    for index, (department_name, department_data) in enumerate(department_totals.items(), start=1):
        sheet.cell(row=row_num, column=1, value=index)  # Sl.No
        sheet.cell(row=row_num, column=2, value=department_name)  # Department Name
        sheet.cell(row=row_num, column=3, value=department_data.get('total_employees', 0))
        sheet.cell(row=row_num, column=4, value=department_data.get('total_per_day_salary', 0))
        sheet.cell(row=row_num, column=5, value=department_data.get('total_paid_days', 0))
        sheet.cell(row=row_num, column=6, value=department_data.get('total_ot_hours', 0))
        sheet.cell(row=row_num, column=7, value=department_data.get('shift_wages', 0))
        sheet.cell(row=row_num, column=8, value=department_data.get('ot_amount', 0))
        sheet.cell(row=row_num, column=9, value=department_data.get('total_rent', 0))
        sheet.cell(row=row_num, column=10, value=department_data.get('total_incentives', 0))
        sheet.cell(row=row_num, column=11, value=department_data.get('total_gross_amount', 0))
        sheet.cell(row=row_num, column=12, value=department_data.get('advance_total', 0))
        sheet.cell(row=row_num, column=13, value=department_data.get('mess_total', 0))
        sheet.cell(row=row_num, column=14, value=department_data.get('store_total', 0))
        sheet.cell(row=row_num, column=15, value=department_data.get('others_total', 0))
        sheet.cell(row=row_num, column=16, value=department_data.get('net_salary_total', 0))
        row_num += 1

    # Add overall totals as the last row
    overall_row = [
        "", "Total", overall_totals.get('total_employees', 0), overall_totals.get('total_per_day_salary', 0),
        overall_totals.get('total_days', 0), overall_totals.get('total_ot_hours', 0),
        overall_totals.get('total_shift_wages', 0), overall_totals.get('total_ot_amount', 0),
        overall_totals.get('total_rent', 0), overall_totals.get('total_incentives', 0),
        overall_totals.get('total_gross_amount', 0), overall_totals.get('total_advance_deductions', 0),
        overall_totals.get('total_mess_deductions', 0), overall_totals.get('total_store_deductions', 0),
        overall_totals.get('total_other_deductions', 0), overall_totals.get('net_salary_total', 0)
    ]
    for col_num, value in enumerate(overall_row, 1):
        sheet.cell(row=row_num, column=col_num, value=value)

    # Prepare response for download
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename=category_totals.xlsx'
    workbook.save(response)
    return response


## all are  class based views we use becaus thats more easy and do a good workin codes i think

# views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from .models import Leader, LeaderCommissionWages
from .forms import LeaderForm, LeaderCommissionWagesForm





# Leader Views
class LeaderListView(ListView):
    model = Leader
    template_name = 'employee/leader/leader_list.html'
    context_object_name = 'leaders'
# views.py
from django.http import HttpResponse
from django.views.generic import DetailView
from openpyxl import Workbook
from .models import Leader, Employee
# views.py
from django.http import HttpResponse
from django.views.generic import DetailView, ListView
from openpyxl import Workbook
from .models import Leader, Employee

class LeaderDetailView(DetailView):
    model = Leader
    template_name = 'employee/leader/leader_detail.html'
    context_object_name = 'leader'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['employees'] = Employee.objects.filter(leader=self.object)
        return context

def export_leader_to_excel(request, pk):
    leader = Leader.objects.get(pk=pk)
    employees = Employee.objects.filter(leader=leader)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"{leader.name} - Employee Details"

    headers = ['Employee Name', 'Position', 'Department', 'Joining Date']
    sheet.append(headers)

    for employee in employees:
        sheet.append([
            employee.employee_name,
            employee.category.name,
            employee.department.name,
            employee.date_of_joining
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{leader.name}_Employee_Details.xlsx"'
    workbook.save(response)
    return response



class LeaderCreateView(CreateView):
    model = Leader
    form_class = LeaderForm
    template_name = 'employee/leader/leader_form.html'
    success_url = reverse_lazy('leader_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_title'] = 'Add Leader'
        return context

class LeaderUpdateView(UpdateView):
    model = Leader
    form_class = LeaderForm
    template_name = 'employee/leader/leader_edit_form.html'
    success_url = reverse_lazy('leader_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_title'] = 'Edit Leader'
        return context

class LeaderDeleteView(DeleteView):
    model = Leader
    template_name = 'employee/leader/leader_confirm_delete.html'
    success_url = reverse_lazy('leader_list')

# LeaderCommissionWages Views
class LeaderCommissionWagesListView(ListView):
    model = LeaderCommissionWages
    template_name = 'employee/leader/leader_commision/commission_wages_list.html'
    context_object_name = 'commission_wages'

class LeaderCommissionWagesCreateView(CreateView):
    model = LeaderCommissionWages
    form_class = LeaderCommissionWagesForm
    template_name = 'employee/leader/leader_commision/commission_wages_form.html'
    success_url = reverse_lazy('commission_wages_list')

class LeaderCommissionWagesUpdateView(UpdateView):
    model = LeaderCommissionWages
    form_class = LeaderCommissionWagesForm
    template_name = 'employee/leader/leader_commision/commission_wages_edit_form.html'
    success_url = reverse_lazy('commission_wages_list')

class LeaderCommissionWagesDeleteView(DeleteView):
    model = LeaderCommissionWages
    template_name = 'employee/leader/leader_commision/commission_wages_confirm_delete.html'
    success_url = reverse_lazy('commission_wages_list')



from django.shortcuts import render
from django.views import View
from .forms import CommissionAttendanceForm
from attendance.models import Attendance
from audit.models import  AuditAttendance
from django.utils import timezone

from django.http import HttpResponse
from openpyxl import Workbook
from django.views import View
from decimal import Decimal
# views.py

from django.views import View
from django.shortcuts import render
from django.http import HttpResponse
from openpyxl import Workbook
from .forms import CommissionAttendanceForm
from decimal import Decimal
from django.utils import timezone
from django.http import HttpResponse
from django.views import View
from openpyxl import Workbook
from io import BytesIO
from decimal import Decimal
from .forms import CommissionAttendanceForm
from django.utils import timezone


from django.shortcuts import render
from django.http import HttpResponse
from django.utils import timezone
from django.views import View
from openpyxl import Workbook
from io import BytesIO
from decimal import Decimal
import calendar

from .forms import CommissionAttendanceForm

class CommissionAttendanceView(View):
    template_name = 'employee/leader/leader_commision/commission_attendance.html'

    def get(self, request):
        form = CommissionAttendanceForm()
        return render(request, self.template_name, {'form': form})

    def post(self, request):
        form = CommissionAttendanceForm(request.POST)
        if form.is_valid():
            leader = form.cleaned_data['leader']
            month = form.cleaned_data['month']

            # Calculate the start and end date of the selected month
            start_date = month.replace(day=1)
            _, last_day = calendar.monthrange(month.year, month.month)
            end_date = month.replace(day=last_day)

            # Get the commission amount for the leader for the selected month
            commission_record = LeaderCommissionWages.objects.filter(
                leader=leader,
                from_date__lte=start_date,
                to_date__gte=end_date
            ).first()
            commission_amount = commission_record.amount if commission_record else Decimal('0.00')

            # Get attendance records for the selected leader's active employees
            employees = Employee.objects.filter(leader=leader, status='active', pf_esi_applicable='yes')

            audit_attendance_list = []
            total_present_days = 0
            total_commission_amount = Decimal('0.00')

            for employee in employees:
                present_days = 0

                # Only check for employees where PF/ESI is applicable
                if employee.pf_esi_applicable == 'yes':
                    attendance_records = AuditAttendance.objects.filter(
                        employee=employee,
                        date__range=[start_date, end_date]
                    )
                    
                    # Log each employee's attendance records
                    print(f"Employee {employee.employee_code} attendance records: {attendance_records}")
                    
                    # Count only 'P' days and log the count
                    present_days = attendance_records.filter(absent='P').count()
                    print(f"Employee {employee.employee_code} - Present Days: {present_days}")

                # Calculate the commission amount for this employee
                employee_commission_amount = present_days * commission_amount

                # Build the attendance record
                audit_attendance = {
                    'employee_code': employee.employee_code,
                    'employee_name': employee.employee_name,
                    'present_days': present_days,
                    'pf_applicable': employee.pf_esi_applicable,
                    'commission_amount': employee_commission_amount
                }
                audit_attendance_list.append(audit_attendance)

                total_present_days += present_days
                total_commission_amount += employee_commission_amount

            # Check for download request
            if 'download' in request.POST:
                return self.generate_excel(audit_attendance_list, total_present_days, total_commission_amount, leader)

            # Render the template with calculated data
            return render(request, self.template_name, {
                'form': form,
                'audit_attendance_list': audit_attendance_list,
                'total_present_days': total_present_days,
                'total_commission_amount': total_commission_amount
            })

        # Render the form again if invalid
        return render(request, self.template_name, {'form': form})

    def generate_excel(self, attendance_list, total_present_days, total_commission_amount, leader):
        # Create a new Excel workbook and add data
        wb = Workbook()
        ws = wb.active
        ws.title = "Commission Attendance"

        # Define headers and leader name row
        headers = ['Leader Name', 'Employee Code', 'Employee Name', 'Present Days', 'Commission Amount']
        ws.append(headers)
        ws.append([leader.name, '', '', '', ''])  # Leader's name in the first cell, others empty

        # Populate attendance data
        for attendance in attendance_list:
            row = [
                '',  # Leave the Leader's Name column empty for data rows
                attendance['employee_code'],
                attendance['employee_name'],
                attendance['present_days'],
                attendance['commission_amount']
            ]
            ws.append(row)

        # Append totals at the end
        ws.append(['Total', '', '', total_present_days, total_commission_amount])

        # Save the workbook to a BytesIO object
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Create HTTP response with Excel data
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=commission_attendance.xlsx'
        return response




from django.shortcuts import render
from django.http import HttpResponse
from django.utils import timezone
from django.views import View
from openpyxl import Workbook
from io import BytesIO
from decimal import Decimal
import calendar

from .forms import CommissionAttendanceForm
from django.shortcuts import render, redirect
from django.views import View
from django.http import HttpResponse
from django.contrib import messages
from openpyxl import Workbook
from io import BytesIO
import calendar
from .forms import CommissionAttendanceForm
from decimal import Decimal

class NonPFCommissionAttendanceView(View):
    template_name = 'employee/leader/leader_commision/commission_non_pf_attendance.html'

    def get(self, request):
        form = CommissionAttendanceForm()
        return render(request, self.template_name, {'form': form})

    def post(self, request):
        form = CommissionAttendanceForm(request.POST)
        if form.is_valid():
            leader = form.cleaned_data['leader']
            month = form.cleaned_data['month']
            start_date = month.replace(day=1)
            _, last_day = calendar.monthrange(month.year, month.month)
            end_date = month.replace(day=last_day)

            # Fetch leader commission amount for the month
            commission_record = LeaderCommissionWages.objects.filter(
                leader=leader,
                from_date__lte=start_date,
                to_date__gte=end_date
            ).first()
            commission_amount = commission_record.amount if commission_record else Decimal('0.00')

            # Query active non-PF employees under this leader
            employees = Employee.objects.filter(leader=leader, status='active', pf_esi_applicable='no')

            # Initialize totals and attendance list
            attendance_list = []
            total_present_days = 0
            total_commission_amount = Decimal('0.00')

            for employee in employees:
                attendance_records = Attendance.objects.filter(
                    employee=employee,
                    date__range=[start_date, end_date]
                )
                present_days = attendance_records.filter(absent='P').count()
                employee_commission_amount = present_days * commission_amount

                # Add to attendance data list
                attendance_list.append({
                    'employee_code': employee.employee_code,
                    'employee_name': employee.employee_name,
                    'present_days': present_days,
                    'commission_amount': employee_commission_amount,
                })

                total_present_days += present_days
                total_commission_amount += employee_commission_amount

            if 'download' in request.POST:
                return self.generate_excel(attendance_list, total_present_days, total_commission_amount, leader)

            return render(request, self.template_name, {
                'form': form,
                'attendance_list': attendance_list,
                'total_present_days': total_present_days,
                'total_commission_amount': total_commission_amount,
            })

        return render(request, self.template_name, {'form': form})

    def generate_excel(self, attendance_list, total_present_days, total_commission_amount, leader):
        wb = Workbook()
        ws = wb.active
        ws.title = "Commission Attendance"

        # Header row with leader's name and report title
        ws.append([f"Commission Attendance Report for {leader.name}"])
        ws.append(["", "", "", "", ""])  # Empty row for spacing
        headers = ['Employee Code', 'Employee Name', 'Present Days', 'Commission Amount']
        ws.append(headers)

        # Populate attendance records
        for attendance in attendance_list:
            ws.append([
                attendance['employee_code'],
                attendance['employee_name'],
                attendance['present_days'],
                float(attendance['commission_amount']),  # Convert to float for better Excel formatting
            ])

        # Totals row
        ws.append(['Total', '', total_present_days, float(total_commission_amount)])

        # Output the workbook to an in-memory file
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Set up HTTP response with Excel download
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{leader.name}_commission_attendance.xlsx"'
        return response





from django.shortcuts import render, redirect
from django.views import View
import openpyxl
from django.http import HttpResponse
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.urls import reverse
from django.views import View
from datetime import datetime, timedelta
from decimal import Decimal
from .forms import ExtraDaysCommissionForm
class ExtraDaysCommissionView(View):
    template_name = 'employee/leader/leader_commision/extra_days_commission.html'

    def get(self, request):
        """Handle GET requests to display the commission form."""
        form = ExtraDaysCommissionForm()
        return render(request, self.template_name, {'form': form})

    def post(self, request):
        """Handle POST requests for commission calculations and Excel export."""
        form = ExtraDaysCommissionForm(request.POST)
        commission_details = []
        
        # Initialize totals
        total_extra_days = 0
        total_commission_amount = Decimal('0.00')

        if form.is_valid():
            leader = form.cleaned_data['leader']
            selected_date = form.cleaned_data['month']
            year = selected_date.year
            month = selected_date.month
            
            # Calculate start and end of the month
            first_day_of_month = datetime(year, month, 1)
            last_day_of_month = (first_day_of_month.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

            # Get the leader's commission rate
            commission_record = LeaderCommissionWages.objects.filter(
                leader=leader,
                from_date__lte=first_day_of_month,
                to_date__gte=last_day_of_month
            ).first()
            
            commission_rate = commission_record.amount if commission_record else Decimal('0.00')
            
            # Retrieve active employees under the selected leader
            employees = Employee.objects.filter(leader=leader, status='active', pf_esi_applicable='yes')
            
            for employee in employees:
                attendance_records = Attendance.objects.filter(
                    employee=employee,
                    date__range=[first_day_of_month, last_day_of_month]
                )
                
                # Calculate extra days based on attendance on weekly off days
                extra_days = sum(
                    1 for record in attendance_records 
                    if record.absent == 'P' and record.date.strftime('%A') == employee.weekly_off
                )

                # Calculate commission amount for extra days
                commission_amount = extra_days * commission_rate
                
                # Add to totals
                total_extra_days += extra_days
                total_commission_amount += commission_amount
                
                # Append employee details
                commission_details.append({
                    'employee_code': employee.employee_code,
                    'employee_name': employee.employee_name,
                    'extra_days': extra_days,
                    'commission_amount': float(commission_amount),
                })

            # Check if the download button was pressed
            if 'download' in request.POST:
                return self.export_to_excel(commission_details, total_extra_days, total_commission_amount, leader.name, selected_date)

            # Render the template with totals and details
            return render(request, self.template_name, {
                'form': form,
                'commission_details': commission_details,
                'total_extra_days': total_extra_days,
                'total_commission_amount': float(total_commission_amount),
            })

        return render(request, self.template_name, {'form': form})

    def export_to_excel(self, commission_details, total_extra_days, total_commission_amount, leader_name, month):
        """Export commission details to an Excel file."""
        # Create a new Excel workbook and add a worksheet
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Extra Days Commission"

        # Add title row
        title_row = f"Extra Days Commission Report for {leader_name} - {month.strftime('%B %Y')}"
        title_cell = worksheet.cell(row=1, column=1, value=title_row)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Merge title cell across the columns (assuming 4 columns)
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

        # Add headers
        headers = ["Employee Code", "Employee Name", "Extra Days", "Commission Amount"]
        for col_num, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=2, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Add data rows
        for row_num, detail in enumerate(commission_details, start=3):
            worksheet.cell(row=row_num, column=1, value=detail['employee_code'])
            worksheet.cell(row=row_num, column=2, value=detail['employee_name'])
            worksheet.cell(row=row_num, column=3, value=detail['extra_days'])
            worksheet.cell(row=row_num, column=4, value=detail['commission_amount'])

        # Add totals row
        totals_row = len(commission_details) + 3
        worksheet.cell(row=totals_row, column=2, value="Total").font = Font(bold=True)
        worksheet.cell(row=totals_row, column=3, value=total_extra_days)
        worksheet.cell(row=totals_row, column=4, value=total_commission_amount)

        # Set response for downloading the file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="extra_days_commission.xlsx"'
        workbook.save(response)
        return response





# views.py
import openpyxl
from django.shortcuts import render, redirect
from django.contrib import messages
from django.views import View
from .models import Leader, LeaderCommissionWages
from .forms import LeaderUploadForm, LeaderCommissionWagesUploadForm
from decimal import Decimal
from django.utils import timezone

class LeaderUploadView(View):
    template_name = 'employee/leader/leader_upload.html'

    def get(self, request):
        form = LeaderUploadForm()
        return render(request, self.template_name, {'form': form})

    def post(self, request):
        form = LeaderUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = form.cleaned_data['file']
            try:
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    name, joining_date, relieving_date, status = row
                    Leader.objects.update_or_create(
                        name=name,
                        defaults={
                            'joining_date': joining_date or timezone.now(),
                            'relieving_date': relieving_date,
                            'status': status.lower(),
                        }
                    )
                messages.success(request, "Leaders uploaded successfully.")
                return redirect('leader_upload')

            except Exception as e:
                messages.error(request, f"Error processing file: {e}")
        
        return render(request, self.template_name, {'form': form})


class LeaderCommissionWagesUploadView(View):
    template_name = 'employee/leader/leader_commission_wages_upload.html'

    def get(self, request):
        form = LeaderCommissionWagesUploadForm()
        return render(request, self.template_name, {'form': form})

    def post(self, request):
        form = LeaderCommissionWagesUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = form.cleaned_data['file']
            try:
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    leader_name, amount, from_date, to_date = row
                    leader = Leader.objects.get(name=leader_name)

                    LeaderCommissionWages.objects.update_or_create(
                        leader=leader,
                        from_date=from_date,
                        to_date=to_date,
                        defaults={'amount': Decimal(amount)},
                    )
                messages.success(request, "Leader Commission Wages uploaded successfully.")
                return redirect('leader_commission_wages_upload')

            except Leader.DoesNotExist:
                messages.error(request, f"Leader '{leader_name}' does not exist.")
            except Exception as e:
                messages.error(request, f"Error processing file: {e}")

        return render(request, self.template_name, {'form': form})


from django.contrib import messages  # Import messages framework
import openpyxl
from django.shortcuts import render, redirect
from django.views import View
from .models import Employee, Leader
from .forms import EmployeeLeaderUploadForm

class EmployeeLeaderUploadView(View):
    template_name = 'employee/employee_leader_upload.html'

    def get(self, request):
        form = EmployeeLeaderUploadForm()
        return render(request, self.template_name, {'form': form})

    def post(self, request):
        form = EmployeeLeaderUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = form.cleaned_data['file']
            try:
                # Load the Excel file
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active
                errors = []
                
                # Iterate over the rows, starting from row 2 (to skip headers)
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    employee_code, leader_name = row
                    
                    # Validate leader existence
                    try:
                        leader = Leader.objects.get(name=leader_name)
                    except Leader.DoesNotExist:
                        errors.append(f"Leader '{leader_name}' does not exist. Skipping employee with code {employee_code}.")
                        continue

                    # Create or update Employee record
                    Employee.objects.update_or_create(
                        employee_code=employee_code,
                        defaults={'leader': leader},
                    )

                # Add messages based on results
                if errors:
                    for error in errors:
                        messages.error(request, error)
                if not errors:
                    messages.success(request, "All employees were uploaded are connected to Leaders  successfully.")
                else:
                    messages.warning(request, "Some employees were skipped due to missing leaders.")

                return render(request, self.template_name, {'form': form})  # Re-render to display messages

            except Exception as e:
                messages.error(request, f"Error processing file: {e}")
        
        return render(request, self.template_name, {'form': form})

from django.shortcuts import render, redirect
from django.urls import reverse
from decimal import Decimal
from decimal import Decimal

def non_pf_salary_department_view(request):
    form = NonPfSalaryForm(request.POST or None)
    department_totals = {}
    overall_totals = {
        'total_p_days': Decimal(0),
        'total_ot_hours': Decimal(0),
        'total_wages': Decimal(0),
        'ot_amount': Decimal(0),
        'gross_amount': Decimal(0),
        'total_deductions': Decimal(0),
        'total_incentives': Decimal(0),
        'net_salary': Decimal(0),
    }

    if form.is_valid():
        selected_date = form.cleaned_data['month']
        year = selected_date.year
        month = selected_date.month

        first_day_of_month = datetime(year, month, 1)
        last_day_of_month = (first_day_of_month.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

        # Filter attendance records
        attendance_records = Attendance.objects.filter(
            date__range=[first_day_of_month, last_day_of_month],
            employee__pf_esi_applicable='no',
            employee__status='active'
        )

        # Filter by category and department if provided
        if form.cleaned_data.get('category'):
            attendance_records = attendance_records.filter(employee__category=form.cleaned_data['category'])
        if form.cleaned_data.get('department'):
            attendance_records = attendance_records.filter(employee__department=form.cleaned_data['department'])

        # Process records department-wise
        for record in attendance_records:
            employee = record.employee
            department_name = employee.department.name if employee.department else 'No Department'

            if department_name not in department_totals:
                department_totals[department_name] = {
                    'total_p_days': Decimal(0),
                    'total_ot_hours': Decimal(0),
                    'total_wages': Decimal(0),
                    'ot_amount': Decimal(0),
                    'gross_amount': Decimal(0),
                    'total_deductions': Decimal(0),
                    'total_incentives': Decimal(0),
                    'net_salary': Decimal(0),
                }

            # Get wages for the employee
            try:
                wages = Wages.objects.filter(
                    employee=employee,
                    start_date__lte=first_day_of_month,
                    end_date__gte=last_day_of_month
                ).first()
                per_day_salary = Decimal(wages.per_day) if wages else Decimal(0)
                rent = Decimal(wages.rent) if wages else Decimal(0)
            except Wages.DoesNotExist:
                per_day_salary = rent = Decimal(0)

            # Get the category hours for the employee (if available)
            category_hours = Decimal(employee.category.hours) if employee.category else Decimal(8)

            # Calculate total_p_days for the employee (Count 'P' days from attendance records)
            total_p_days = attendance_records.filter(employee=employee, absent='P').count()

            # Calculate total_ot_hours for the employee (Sum of ot_hours from attendance records)
            total_ot_hours = attendance_records.filter(employee=employee).aggregate(Sum('ot_hours'))['ot_hours__sum'] or Decimal(0)

            # Ensure total_ot_hours is a Decimal type before using it in calculations
            total_ot_hours = Decimal(total_ot_hours)

            # Calculate OT amount based on the total OT hours and category hours
            ot_amount = round((per_day_salary / category_hours) * total_ot_hours)

            # Calculate total wages based on the present days
            total_wages = total_p_days * per_day_salary

            # Calculate gross amount (wages + OT amount + rent)
            gross_amount = total_wages + ot_amount + rent + data['incentives'] 

            employee_incentives = Incentive.objects.filter(
                employee=employee,
                date__month=month,
                date__year=year
            ).aggregate(total_incentives=Coalesce(Sum('incentive_amount'), Decimal(0)))

            data['incentives'] = employee_incentives['total_incentives'] or Decimal(0)

            # Deductions
            deductions = Deduction.objects.filter(employee=employee, date__month=month, date__year=year)
            total_deductions = deductions.aggregate(Sum('amount'))['amount__sum'] or Decimal(0)

            # Calculate net salary
            net_salary = gross_amount - total_deductions

            # Update department totals
            department_totals[department_name]['total_p_days'] += total_p_days
            department_totals[department_name]['total_ot_hours'] += total_ot_hours
            department_totals[department_name]['total_wages'] += total_wages
            department_totals[department_name]['ot_amount'] += ot_amount
            department_totals[department_name]['gross_amount'] += gross_amount
            department_totals[department_name]['total_deductions'] += total_deductions
            department_totals[department_name]['total_incentives'] += data['incentives'] 
            department_totals[department_name]['net_salary'] += net_salary

        # Calculate overall totals
        for dept_totals in department_totals.values():
            for key in overall_totals:
                overall_totals[key] += dept_totals[key]

        # Convert Decimal to float for JSON serialization
        department_totals_serializable = {
            dept: {key: float(value) for key, value in totals.items()}
            for dept, totals in department_totals.items()
        }

        overall_totals_serializable = {key: float(value) for key, value in overall_totals.items()}

        # Save the data in the session for export
        request.session.update({
            'department_totals': department_totals_serializable,
            'overall_totals': overall_totals_serializable,
            'month': selected_date.strftime('%Y-%m'),
        })

    return render(request, 'employee/employee_salary/non_pf_salary_department.html', {
        'form': form,
        'department_totals': department_totals_serializable,
        'overall_totals': overall_totals_serializable,
    })






# Helper function to convert Decimal to float
def convert_decimals_to_float(data):
    if isinstance(data, dict):
        return {key: convert_decimals_to_float(value) for key, value in data.items()}
    elif isinstance(data, list):
        return [convert_decimals_to_float(item) for item in data]
    elif isinstance(data, Decimal):
        return float(data)
    return data


from .forms import MonthForm

def extra_days_department_totals_view(request):
    form = MonthForm(request.POST or None)
    salary_details = []
    selected_month_name = None
    
    # Initialize overall totals
    total_paid_days = 0
    total_ot_hours = 0
    all_total_wages = 0
    
    total_gross_amount = 0
    all_total_deductions = 0
    total_incentives = 0
    total_advance_deductions = 0
    total_canteen_deductions = 0
    total_mess_deductions = 0
    total_store_deductions = 0
    total_other_deductions = 0
    total_net_salary = 0
    total_ot_amount = 0
    total_wages = 0
    total_deductions = 0
    total_net_salary = 0
    total_present_days = 0
   
    total_per_day_salary = 0
    total_paid_days = 0
    total_pf_amount = 0
    basic_salary = 0
    da_salary = 0
    hra_salary = 0
    rent=0
     

    # Initialize category and department totals
    category_totals = {}
    department_totals = {}

    # Overall totals for the organization
    overall_totals = {
        'total_employees': 0,
        'total_advance_deductions': 0,
        'total_mess_deductions': 0,
        'total_store_deductions': 0,
        'total_other_deductions': 0,
        'total_deductions': 0,
        'net_salary_total': 0,
        'total_incentives': 0,
        'total_gross_amount': 0,
        'total_days':0,
        'all_employees':0,
        'total_per_day_salary':0,
        'total_shift_wages':0,
        'total_basic_amount':0,
        'total_da_amount':0,
        'total_hra_amount':0,
        'total_pf_amount':0,
        'total_esi_amount': 0 ,
        'total_ot_hours':0,
        'total_ot_amount':0,
        'total_rent': 0 

       
       
        

    }

    # Initialize the total employees count for the entire organization
    overall_total_employees = 0
    selected_month = None
    selected_year = None
    month = None
    year = None
    month_name = None

    salary_details = []
    
    # Initialize total variables
    total_extra_days_wages = 0
    total_ot_amount = 0
    total_incentives = 0

    total_net_salary = 0
    total_extra_days = 0
    total_ot_hours = 0
    total_per_day_salary = 0
    total_rent = 0  # Add total_rent to accumulate the rent values

    if form.is_valid():
        selected_date = form.cleaned_data['month']
        year = selected_date.year
        month = selected_date.month
        selected_year = selected_date.year
        selected_month = selected_date.month
        month_name = selected_date.strftime('%B')


        first_day_of_month = datetime(year, month, 1)
        last_day_of_month = (first_day_of_month.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

        employees = Employee.objects.filter(pf_esi_applicable='yes', status='active')

        # Filter employees based on form data
        if form.cleaned_data.get('category'):
            employees = employees.filter(category=form.cleaned_data['category'])
        if form.cleaned_data.get('department'):
            employees = employees.filter(department=form.cleaned_data['department'])

        # Get attendance records for all employees in the selected date range
        attendance_records = Attendance.objects.filter(
            date__range=[first_day_of_month, last_day_of_month],
            employee__in=employees
        )

        employees_dict = {}
        for employee in employees:
            employees_dict[employee] = {
                'extra_days': 0,
                'total_ot_hours': Decimal(0),
                'incentives': Decimal(0),
                'weekly_off_day': employee.weekly_off,
                'total_present_days': 0,
                'rent':0
            }

        for record in attendance_records:
            employee = record.employee
            weekly_off_day = employees_dict[employee]['weekly_off_day']

            # Check if this weekly off day is also a national or festival holiday
            if record.date.strftime('%A') == weekly_off_day:
                is_holiday = Holiday.objects.filter(
                    Q(date=record.date),
                    Q(holiday_type='NH') | Q(holiday_type='FH')  # Only consider National and Festival Holidays
                ).exists()

                # If it's not a holiday and the employee was present, count it as an extra day
                if not is_holiday and record.absent == 'P':
                    employees_dict[employee]['extra_days'] += 1

            # If the employee was present, count it as a present day
            if record.absent == 'P':
                employees_dict[employee]['total_present_days'] += 1

            # Add overtime hours if applicable
            if record.ot_hours:
                ot_hours = Decimal(record.ot_hours)
                employees_dict[employee]['total_ot_hours'] += ot_hours

        # Calculate wages, incentives, and other details for each employee
        for i, (employee, data) in enumerate(employees_dict.items(), start=1):
            try:
                # Fetch the wages for the employee
                wages = Wages.objects.get(employee=employee, start_date__lte=first_day_of_month, end_date__gte=last_day_of_month)
                per_day_salary = Decimal(wages.per_day)
                rent = Decimal(wages.rent) if wages else Decimal(0)
            except Wages.DoesNotExist:
                per_day_salary = Decimal(0)
                rent = Decimal(0)



            overall_total_employees += 1

            employee_incentives = Incentive.objects.filter(
                employee=employee,
                date__month=month,
                date__year=year
            ).aggregate(total_incentives=Coalesce(Sum('incentive_amount'), Decimal(0)))

            data['incentives'] = employee_incentives['total_incentives'] or Decimal(0)



            
            # Calculate extra days wages
            extra_days_wages = data['extra_days'] * per_day_salary + data['incentives']  + rent
            

            # Get the number of hours in the employee's category
            category_hours = Decimal(employee.category.hours) if employee.category else Decimal(0)
            ot_amount = int(round((per_day_salary / category_hours) * data['total_ot_hours'] * 1))
            
            # Calculate net salary
            net_salary = extra_days_wages + ot_amount  
            net_salary = round(net_salary)  # Round net salary to the nearest 10

            # Update total variables for the session
            total_extra_days_wages += float(extra_days_wages)
            total_ot_amount += float(ot_amount)
            total_incentives += data['incentives']
            total_net_salary += float(net_salary)
            

            # Accumulate extra days, total OT hours, and per day salary
            total_extra_days += data['extra_days']
            total_ot_hours += data['total_ot_hours']
            total_per_day_salary += float(per_day_salary)

            # Append salary details for each employee
            salary_details.append({
                'sl_no': i,
                'employee_code': employee.employee_code,
                'employee_name': employee.employee_name,
                'name': employee.employee_name,
                'department': employee.department.name if employee.department else '',
                'designation': employee.designation.name if employee.designation else '',
                'extra_days': data['extra_days'],
                'total_ot_hours': data['total_ot_hours'],
                'per_day_salary': float(per_day_salary),
                'rent': float(rent),  # Include rent in details
                'extra_days_wages': float(extra_days_wages),
                'ot_amount': float(ot_amount),
                'total_incentives': float(data['incentives']),
                'net_salary': float(net_salary),
            })


            # Update overall totals
            total_paid_days += data['extra_days']
            total_ot_hours += data['total_ot_hours']
            all_total_wages += total_wages
            
            
            all_total_deductions += total_deductions
            
            
            total_net_salary += net_salary  # Accumulate net salary here
       
            
        
            # Accumulate department totals (new logic)
            department_name = employee.department.name if employee.department else 'Uncategorized'
            if department_name not in department_totals:
                department_totals[department_name] = {
                    'total_per_day_salary': 0,  # Track total per day salary
                    'total_employees': 0,  # Track the total number of employees
                    'shift_wages':0, 
             
                    'ot_amount': 0,  # Track total advance deductions
                    
                    'total_rent': 0,  # Track total mess deductions
                    'store_total': 0,  # Track total store deductions
                    'others_total': 0,  # Track total other deductions
                    'total_incentives': 0,  # Track total deductions in the department
                    'net_salary_total': 0,
                    
                    'total_gross_amount': 0,
                    'total_paid_days': 0,  # Ensure this key is initialized
                    'total_rent': 0, 
                    'total_ot_hours':0,
        
                }
            
            # Update department totals for this employee
      
            department_totals[department_name]['total_ot_hours'] += data['total_ot_hours']
            department_totals[department_name]['ot_amount'] +=  ot_amount 
            department_totals[department_name]['total_rent'] +=  rent
        
            
            department_totals[department_name]['total_paid_days'] += data['extra_days']
            department_totals[department_name]['total_employees'] += 1
            department_totals[department_name]['total_incentives'] += data['incentives']
   
          
         
            department_totals[department_name]['net_salary_total'] += net_salary
         
            department_totals[department_name]['total_gross_amount'] += extra_days_wages
 

            # Update overall totals for the entire organization
            overall_totals['total_rent'] +=  rent
            overall_totals['total_ot_amount'] +=  ot_amount 
            overall_totals['total_incentives'] += data['incentives']
            overall_totals['total_ot_hours']  +=  data['total_ot_hours']
           
            overall_totals['total_shift_wages']  += total_wages
            overall_totals['total_days']  += data['extra_days']

            overall_totals['net_salary_total'] += net_salary
        
            overall_totals['total_gross_amount'] += extra_days_wages

        # Update the overall employee count
            overall_totals['total_employees'] = overall_total_employees

    department_totals = dict(sorted(department_totals.items()))

    request.session['department_totals'] = convert_decimals_to_float(department_totals)
    request.session['overall_totals'] = convert_decimals_to_float(overall_totals)
    request.session['selected_month'] = month_name  # Store the selected month
    request.session['selected_year'] = year 
    # Render the response
    return render(request, 'employee/employee_salary/extra_days_department_salary.html', {
        'form': form,
        'salary_details': salary_details,
        'category_totals': category_totals,
        'department_totals': department_totals,
        'overall_totals': overall_totals,
        'overall_total_employees': overall_total_employees,
        'selected_month': month_name,  # Pass selected month
    'selected_year': year,  # Pass selected year
        
    })

from django.shortcuts import redirect
from django.contrib import messages
from .models import ExtraDaysAbstract

def save_extra_days_abstract(request):
    if request.method == "POST":
        # Retrieve data from the session
        department_totals = request.session.get('department_totals', {})
        selected_month = request.session.get('selected_month', 'Unknown')
        selected_year = request.session.get('selected_year', 'Unknown')

        # Ensure year is an integer
        try:
            selected_year = int(selected_year)
        except ValueError:
            messages.error(request, "Invalid year format.")
            return redirect('extra_days_department_salary')  # Replace with your view name

        # Iterate over departments and save or update records
        for department_name, data in department_totals.items():
            obj, created = ExtraDaysAbstract.objects.update_or_create(
                month=selected_month,
                year=selected_year,
                department_name=department_name,
                defaults={
                    'total_employees': data.get('total_employees', 0),
                    'extra_days_worked': data.get('total_paid_days', 0),
                    'incentives': data.get('total_incentives', 0.0),
                    'rent': data.get('total_rent', 0.0),
                    'total_gross_amount': data.get('total_gross_amount', 0.0),
                    'ot_hours': data.get('total_ot_hours', 0),
                    'ot_amount': data.get('ot_amount', 0.0),
                    'net_salary': data.get('net_salary_total', 0.0),
                }
            )
            if created:
                messages.success(request, f"Created record for {department_name} ({selected_month} {selected_year}).")
            else:
                messages.info(request, f"Updated record for {department_name} ({selected_month} {selected_year}).")

        # Redirect to the appropriate page
        messages.success(request, "All department records have been successfully saved.")
        return redirect('extra_days_department_totals_view')  # Replace with your view name
    else:
        # If accessed via GET, redirect to another page
        return redirect('extra_days_department_totals_view')  # Replace with your view name


import openpyxl
from openpyxl.styles import Alignment, Font
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Alignment, Font
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse

def extra_days_export_department_totals_excel(request):
    # Initialize workbook and worksheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Department Totals"
    
    # Set headers
    headers = [
        "Sl.No", "Department", "Total Employees", "Extra Days Worked",
        "Incentives", "Rent", "Total Gross Amount", "OT Hours",
        "OT Amount", "Net Salary"
    ]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Retrieve data from session or fallback
    department_totals = request.session.get('department_totals', {})
    overall_totals = request.session.get('overall_totals', {})
    selected_month = request.session.get('selected_month', 'Unknown Month')
    selected_year = request.session.get('selected_year', 'Unknown Year')

    # Write department data
    row_num = 2
    for index, (department_name, department_data) in enumerate(department_totals.items(), start=1):
        sheet.cell(row=row_num, column=1, value=index)  # Sl.No
        sheet.cell(row=row_num, column=2, value=department_name)  # Department Name
        sheet.cell(row=row_num, column=3, value=department_data.get('total_employees', 0))  # Total Employees
        sheet.cell(row=row_num, column=4, value=department_data.get('total_paid_days', 0))  # Extra Days Worked
        sheet.cell(row=row_num, column=5, value=department_data.get('total_incentives', 0))  # Incentives
        sheet.cell(row=row_num, column=6, value=department_data.get('total_rent', 0))  # Rent
        sheet.cell(row=row_num, column=7, value=department_data.get('total_gross_amount', 0))  # Total Gross Amount
        sheet.cell(row=row_num, column=8, value=department_data.get('total_ot_hours', 0))  # OT Hours
        sheet.cell(row=row_num, column=9, value=department_data.get('ot_amount', 0))  # OT Amount
        sheet.cell(row=row_num, column=10, value=department_data.get('net_salary_total', 0))  # Net Salary
        row_num += 1

    # Write overall totals as the last row
    overall_row = [
        "", "Total", overall_totals.get('total_employees', 0),
        overall_totals.get('total_days', 0), overall_totals.get('total_incentives', 0),
        overall_totals.get('total_rent', 0), overall_totals.get('total_gross_amount', 0),
        overall_totals.get('total_ot_hours', 0), overall_totals.get('total_ot_amount', 0),
        overall_totals.get('net_salary_total', 0)
    ]
    for col_num, value in enumerate(overall_row, 1):
        cell = sheet.cell(row=row_num, column=col_num, value=value)
        cell.font = Font(bold=True)  # Highlight totals
        cell.alignment = Alignment(horizontal='center')

    # Prepare response for download
    filename = f"Extra_Days_abstract_{selected_month}_{selected_year}.xlsx"
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename={filename}'
    workbook.save(response)
    return response




from django.shortcuts import render
from .models import ExtraDaysAbstract, NonPFAbstract
def reports_list(request):
  
    return render(request, 'employee/reports/reports.html',)

# View for ExtraDaysAbstract
def extradays_abstract_list(request):
    unique_months_years = ExtraDaysAbstract.objects.values('month', 'year').distinct()
    return render(request, 'employee/reports/extradays_abstract_list.html', {'unique_months_years': unique_months_years})
from django.shortcuts import render
from .models import ExtraDaysAbstract

def extradays_abstract_details(request, month, year):
    records = ExtraDaysAbstract.objects.filter(month=month, year=year)
    
    # Initialize totals
    total_employees = 0
    total_extra_days_worked = 0
    total_incentives = 0
    total_rent = 0
    total_gross_amount = 0
    total_ot_hours = 0
    total_ot_amount = 0
    total_net_salary = 0

    # Loop through records and calculate totals
    for record in records:
        total_employees += record.total_employees
        total_extra_days_worked += record.extra_days_worked
        total_incentives += record.incentives
        total_rent += record.rent
        total_gross_amount += record.total_gross_amount
        total_ot_hours += record.ot_hours
        total_ot_amount += record.ot_amount
        total_net_salary += record.net_salary
    
    # Render the template with records and totals
    return render(request, 'employee/reports/extradays_abstract_details.html', {
        'month': month,
        'year': year,
        'records': records,
        'total_employees': total_employees,
        'total_extra_days_worked': total_extra_days_worked,
        'total_incentives': total_incentives,
        'total_rent': total_rent,
        'total_gross_amount': total_gross_amount,
        'total_ot_hours': total_ot_hours,
        'total_ot_amount': total_ot_amount,
        'total_net_salary': total_net_salary
    })


# View for NonPFAbstract
def nonpf_abstract_list(request):
    unique_months_years = NonPFAbstract.objects.values('month', 'year').distinct()
    return render(request, 'employee/reports/nonpf_abstract_list.html', {'unique_months_years': unique_months_years})
from django.shortcuts import render
from .models import NonPFAbstract

def nonpf_abstract_details(request, month, year):
    records = NonPFAbstract.objects.filter(month=month, year=year)
    
    # Initialize totals
    total_employees = 0
    total_paid_days = 0
    total_ot_hours = 0
    total_ot_amount = 0
    total_shift_wages = 0
    total_per_day_salary = 0
    total_rent = 0
    total_advance_total = 0
    total_canteen_total = 0
    total_mess_total = 0
    total_store_total = 0
    total_others_total = 0
    total_deductions = 0
    total_gross_amount = 0
    total_net_salary = 0
    total_incentives = 0

    # Loop through records and calculate totals
    for record in records:
        total_employees += record.total_employees
        total_paid_days += record.total_paid_days
        total_ot_hours += record.total_ot_hours
        total_ot_amount += record.ot_amount
        total_shift_wages += record.shift_wages
        total_per_day_salary += record.per_day_salary
        total_rent += record.rent
        total_advance_total += record.advance_total
        total_canteen_total += record.canteen_total
        total_mess_total += record.mess_total
        total_store_total += record.store_total
        total_others_total += record.others_total
        total_deductions += record.total_deductions
        total_gross_amount += record.total_gross_amount
        total_net_salary += record.net_salary
        total_incentives += record.total_incentives
    
    # Render the template with records and totals
    return render(request, 'employee/reports/nonpf_abstract_details.html', {
        'month': month,
        'year': year,
        'records': records,
        'total_employees': total_employees,
        'total_paid_days': total_paid_days,
        'total_ot_hours': total_ot_hours,
        'total_ot_amount': total_ot_amount,
        'total_shift_wages': total_shift_wages,
        'total_per_day_salary': total_per_day_salary,
        'total_rent': total_rent,
        'total_advance_total': total_advance_total,
        'total_canteen_total': total_canteen_total,
        'total_mess_total': total_mess_total,
        'total_store_total': total_store_total,
        'total_others_total': total_others_total,
        'total_deductions': total_deductions,
        'total_gross_amount': total_gross_amount,
        'total_net_salary': total_net_salary,
        'total_incentives': total_incentives
    })



from django.shortcuts import render
from audit.models import PFAbstract

# List View for PFAbstract
def pfabstract_list(request):
    # Get distinct month and year combinations
    unique_months_years = PFAbstract.objects.values('month', 'year').distinct()
    return render(request, 'employee/reports/pfabstract_list.html', {'unique_months_years': unique_months_years})

# Detail View for PFAbstract
def pfabstract_details(request, month, year):
    # Get records for the specified month and year
    records = PFAbstract.objects.filter(month=month, year=year)
    
    # Calculate totals
    totals = {
        'total_employees': sum(record.total_employees for record in records),
        'shift_wages': sum(record.shift_wages for record in records),
        'total_per_day_salary': sum(record.total_per_day_salary for record in records),
        'total_rent': sum(record.total_rent for record in records),
        'total_ot_hours': sum(record.total_ot_hours for record in records),
        'total_basic_amount': sum(record.total_basic_amount for record in records),
        'total_da_amount': sum(record.total_da_amount for record in records),
        'total_hra_amount': sum(record.total_hra_amount for record in records),
        'total_pf_amount': sum(record.total_pf_amount for record in records),
        'total_esi_amount': sum(record.total_esi_amount for record in records),
        'total_paid_days': sum(record.total_paid_days for record in records),
        'advance_total': sum(record.advance_total for record in records),
        'mess_total': sum(record.mess_total for record in records),
        'store_total': sum(record.store_total for record in records),
        'others_total': sum(record.others_total for record in records),
        'total_deductions': sum(record.total_deductions for record in records),
        'net_salary_total': sum(record.net_salary_total for record in records),
        'total_gross_amount': sum(record.total_gross_amount for record in records),
    }
    
    return render(request, 'employee/reports/pfabstract_details.html', {
        'month': month,
        'year': year,
        'records': records,
        'totals': totals,
    })
