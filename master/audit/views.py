from django.urls import reverse
import datetime
from datetime import datetime
from django.shortcuts import render, get_object_or_404, redirect
from employee.models import Employee
from .forms import AuditAdjustmentDateForm, AuditAdjustmentEntryForm, AuditAdjustmentReportViewForm
from datetime import datetime, timedelta
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from collections import defaultdict
from .forms import TimeCardForm
from datetime import timedelta
from django.utils.dateformat import DateFormat
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from .forms import MonthlyAttendanceForm, PunchingReportForm, PfSalaryForm
from .forms import AbsentListForm

import logging
logger = logging.getLogger(__name__)
from django.db.models import Q


# Create your views here.
def audit_home(request):
    return render(request, 'audit_home.html', )


# Create your views here.
def shift_rotation_view(request):
    return render(request, 'attendance_audit/shift_rotation.html', )






from .models import AuditAttendance
from django.shortcuts import render
from django.contrib import messages
from .forms import AttendanceImportForm

def audit_attendance_list_view(request):
    # Fetch all AuditAttendance records from the database
    audit_attendances = AuditAttendance.objects.all()
    return render(request, 'attendance_audit/audit_attendance_list.html', {'audit_attendances': audit_attendances})

    
from datetime import datetime, time
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import AuditAttendance
from datetime import datetime, time, timedelta
from django.shortcuts import render, redirect
from django.contrib import messages

def audit_import_attendance_view(request):
    if request.method == 'POST':
        form = AttendanceImportForm(request.POST)
        if form.is_valid():
            success, message = form.import_attendance()
            if success:
                # Redirect to the audit attendance list page after successful import
                return redirect('audit_attendance_list')
            else:
                messages.warning(request, message)
    else:
        form = AttendanceImportForm()


    return render(request, 'attendance_audit/import_attendance.html')







from django.shortcuts import render, redirect
from .models import Holiday
from .forms import HolidayForm, EmployeeForm, SinglePunchReportForm

def holiday_create_view(request):
    form = HolidayForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('holiday_list_view')  # Redirect to a list view after saving
    return render(request, 'holiday/holiday_form.html', {'form': form})


def holiday_edit_view(request, pk):
    holiday = get_object_or_404(Holiday, pk=pk)
    form = HolidayForm(request.POST or None, instance=holiday)
    if form.is_valid():
        form.save()
        return redirect('holiday_list_view')  # Redirect to a list view after editing
    return render(request, 'holiday/holiday_form.html', {'form': form, 'holiday': holiday})



def holiday_delete_view(request, pk):
    holiday = get_object_or_404(Holiday, pk=pk)
    if request.method == 'POST':
        holiday.delete()
        return redirect('holiday_list_view')  # Redirect to a list view after deleting
    return render(request, 'holiday/holiday_confirm_delete.html', {'holiday': holiday})

def holiday_list_view(request):
    holidays = Holiday.objects.all()
    return render(request, 'holiday/holiday_list.html', {'holidays': holidays})





from employee.forms import EmployeeEditForm


def audit_employee_list(request):
    employees = Employee.objects.filter(status='active', pf_esi_applicable='yes')
    return render(request, 'audit_employee/employee_list.html', {'employees': employees})


def audit_employee_detail(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    return render(request, 'audit_employee/employee_detail.html', {'employee': employee})

def audit_employee_create(request):
    if request.method == 'POST':
        form = EmployeeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('audit_employee_list')
    else:
        form = EmployeeForm()
    return render(request, 'audit_employee/employee_form.html', {'form': form})

def audit_employee_update(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        form = EmployeeEditForm(request.POST, instance=employee)
        if form.is_valid():
            form.save()
            return redirect('audit_employee_list')
    else:
        form = EmployeeEditForm(instance=employee)
    return render(request, 'audit_employee/employee_edit_form.html', {'form': form})

def audit_employee_delete(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        employee.delete()
        return redirect('audit_employee_list')
    return render(request, 'audit_employee/employee_confirm_delete.html', {'object': employee})

from django.views.generic import ListView

class InactiveEmployeeListView(ListView):
    model = Employee
    template_name = 'audit_employee/inactive_employees.html'
    context_object_name = 'inactive_employees'

    def get_queryset(self):
        return Employee.objects.filter(status='inactive', pf_esi_applicable='yes')
    

from django.shortcuts import render
from .forms import PunchingReportForm

import logging

logger = logging.getLogger(__name__)

def audit_punching_report_view(request):
    if request.method == 'POST':
        form = PunchingReportForm(request.POST)
        if form.is_valid():
            date = form.cleaned_data['date']
            # Filter attendance records where absent is 'P'
            attendance_records = AuditAttendance.objects.filter(date=date, absent='P').order_by()
            logger.debug(f"Retrieved attendance records: {attendance_records}")
    else:
        form = PunchingReportForm()
        attendance_records = None

    return render(request, 'audit_reports/punching_report.html', {'form': form, 'attendance_records': attendance_records})

def audit_single_punch_report_view(request):
    punch_and_absent_list = []
    if request.method == 'POST':
        form = SinglePunchReportForm(request.POST)
        if form.is_valid():
            date = form.cleaned_data['date']

            # Fetch records where either intime or outtime is present and matches the selected date
            punch_and_absent_list = AuditAttendance.objects.filter(date=date).filter(
                Q(intime__isnull=False, outtime__isnull=True) | Q(intime__isnull=True, outtime__isnull=False)
            )

            logger.debug(f"Retrieved punch_and_absent_list: {punch_and_absent_list}")
        else:
            logger.debug(f"Form errors: {form.errors}")
    else:
        form = SinglePunchReportForm()

    context = {
        'form': form,
        'punch_and_absent_list': punch_and_absent_list,
    }
    
    return render(request, 'audit_reports/single_punch_report.html', context)




def audit_absent_list_view(request):
    attendance_absent_records = []
    absent_employees = []
    
    if request.method == 'POST':
        form = AbsentListForm(request.POST)
        if form.is_valid():
            date = form.cleaned_data['date']

            # Fetch attendance records where absent is 'Yes' for the selected date
            attendance_absent_records =AuditAttendance.objects.filter(date=date, absent='AB')
            
            # Get all active employees
            active_employees = Employee.objects.filter(status='active', pf_esi_applicable='yes')
            
            # Get employees who have attendance records for the selected date
            employees_with_attendance = AuditAttendance.objects.filter(date=date).values_list('employee', flat=True)
            
            # Filter out active employees who do not have attendance records for the selected date
            absent_employees = active_employees.exclude(id__in=employees_with_attendance)

            logger.debug(f"Retrieved attendance_absent_records: {attendance_absent_records}")
            logger.debug(f"Retrieved absent_employees: {absent_employees}")
        else:
            logger.debug(f"Form errors: {form.errors}")
    else:
        form = AbsentListForm()

    context = {
        'form': form,
        'attendance_absent_records': attendance_absent_records,
        'absent_employees': absent_employees,
    }
    
    return render(request, 'audit_reports/absent_list.html', context)




def audit_date_selection_view(request):
    if request.method == 'POST':
        form = AuditAdjustmentDateForm(request.POST)
        if form.is_valid():
            date = form.cleaned_data['date']
            return redirect('audit_adjustment_entry_view', date=date.strftime('%Y-%m-%d'))  # Redirect to attendance details view with selected date
    else:
        form = AuditAdjustmentDateForm()

    return render(request, 'audit_entries/date_selection.html', {'form': form})
# At the top of your views.py



def audit_adjustment_attendance_details_view(request, date):
    date_obj = datetime.strptime(date, '%Y-%m-%d').date()

    if request.method == 'POST':
        form = AuditAdjustmentEntryForm(request.POST)
        if form.is_valid():
            employee_code = form.cleaned_data['employee_code']
            absent_status = form.cleaned_data['absent']
            employee = get_object_or_404(Employee, employee_code=employee_code)

            # Check if an attendance record already exists
            attendance_record =AuditAttendance.objects.filter(employee=employee, date=date_obj).first()

            if attendance_record:
                # Update the existing record
                attendance_record.absent = absent_status
                attendance_record.category = employee.category.name if employee.category else ''
                attendance_record.department = employee.department.name if employee.department else ''
                attendance_record.save()
            else:
                # Create a new attendance record
                AuditAttendance.objects.create(
                    employee=employee,
                    date=date_obj,
                    absent=absent_status,
                    category=employee.category.name if employee.category else '',
                    department=employee.department.name if employee.department else ''
                )
            return redirect(reverse('audit_adjustment_entry_view', kwargs={'date': date}))

    else:
        form = AuditAdjustmentDateForm()

    attendance_records = AuditAttendance.objects.filter(date=date_obj)
    active_employees = Employee.objects.filter(status='active', pf_esi_applicable='yes')
    employees_with_attendance = AuditAttendance.objects.filter(date=date_obj).values_list('employee_id', flat=True)
    absent_employees = active_employees.exclude(id__in=employees_with_attendance)

    context = {
        'attendance_records': attendance_records,
        'absent_employees': absent_employees,
        'date_obj': date_obj,
        'form': form,
        'adjustment_form': AuditAdjustmentEntryForm(),
    }
    return render(request, 'audit_entries/attendance_details.html', context)



def audit_adjustment_report_view(request):
    punch_and_absent_list = []
    if request.method == 'POST':
        form = AdjustmentReportViewForm(request.POST)
        if form.is_valid():
            date = form.cleaned_data['date']

            # Fetch records where either intime or outtime is present and matches the selected date
            punch_and_absent_list = AuditAttendance.objects.filter(date=date).filter(
                Q(intime__isnull=False, outtime__isnull=False) |Q(intime__isnull=True, outtime__isnull=False) |Q(intime__isnull=False, outtime__isnull=True)
            ).filter(absent="P")

            logger.debug(f"Retrieved punch_and_absent_list: {punch_and_absent_list}")
        else:
            logger.debug(f"Form errors: {form.errors}")
    else:
        form = SinglePunchReportForm()

    context = {
        'form': form,
        'punch_and_absent_list': punch_and_absent_list,
    }
    
    return render(request, 'audit_reports/adjustment_entry_report.html', context)
from datetime import timedelta
from django.shortcuts import render
from .models import Employee, Holiday, AuditAttendance
from .forms import MonthlyAttendanceForm

def audit_monthly_attendance_view(request):
    # Initialize variables
    attendance_data = []
    dates = []
    start_date = end_date = None

    if request.method == 'POST':
        form = MonthlyAttendanceForm(request.POST)
        if form.is_valid():
            # Extract form data
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            category = form.cleaned_data['category']
            department = form.cleaned_data['department']

            # Generate the date range between start and end dates
            dates = [start_date + timedelta(days=x) for x in range((end_date - start_date).days + 1)]

            # Retrieve active employees based on category and department filters
            employees = Employee.objects.filter(status='active', pf_esi_applicable='yes')
            if category:
                employees = employees.filter(category=category)
            if department:
                employees = employees.filter(department=department)

            # Fetch holiday records within the date range
            holidays = Holiday.objects.filter(date__in=dates)
            holiday_dict = {holiday.date: holiday.holiday_type for holiday in holidays}

            # Process attendance for each employee within the given date range
            for employee in employees:
                employee_attendance = []
                present_days = 0
                absent_days = 0
                nh_count = 0  # National Holiday count
                fh_count = 0  # Festival Holiday count
                total_nh_fh_count = 0
                weekly_off_day = employee.weekly_off

                # Iterate through each date in the date range to calculate attendance
                for date in dates:
                    # Determine the attendance status based on holiday or weekly off
                    if date in holiday_dict:
                        attendance_status = holiday_dict[date]  # NH, FH, or WH
                        if attendance_status == 'NH':
                            nh_count += 1
                        elif attendance_status == 'FH':
                            fh_count += 1
                        total_nh_fh_count = nh_count + fh_count
                    else:
                        # Check if it's a weekly off day
                        if date.strftime('%A') == weekly_off_day:
                            attendance_status = 'WO'  # Weekly Off
                        else:
                            # Retrieve attendance record for the specific date
                            attendance_record = AuditAttendance.objects.filter(employee=employee, date=date).first()
                            if attendance_record:
                                attendance_status = 'A' if attendance_record.absent == 'AB' else attendance_record.absent
                            else:
                                attendance_status = 'A'  # Absent if no record found

                    # Track present and absent days based on attendance status
                    if attendance_status == 'P':
                        present_days += 1
                    elif attendance_status == 'A':
                        absent_days += 1

                    # Append the calculated attendance status to the employee's record
                    employee_attendance.append(attendance_status)
                # Calculate total days as sum of present days and total NH/FH holidays
                total_days = total_nh_fh_count + present_days

                # Aggregate the employee's attendance data
                attendance_data.append({
                    'employee': employee,
                    'attendance': employee_attendance,
                    'present_days': present_days,
                    'absent_days': absent_days,
                    'nh_count': nh_count,  # National Holiday count
                    'fh_count': fh_count,  # Festival Holiday count
                    'total_nh_fh_count': total_nh_fh_count,  # Total holidays (NH + FH)
                    'total_days':total_days
                })
                 # Sort the attendance data by employee_code in ascending order
            attendance_data.sort(key=lambda x: x['employee'].employee_code)

            # Store processed attendance data in the session for potential downloads or reports
            request.session['attendance_data'] = [
                {
                    'employee_code': data['employee'].employee_code,
                    'employee_name': data['employee'].employee_name,
                    'category': data['employee'].category.name if data['employee'].category else '',
                    'department': data['employee'].department.name if data['employee'].department else '',
                    'attendance': data['attendance'],
                    'present_days': data['present_days'],
                    'absent_days': data['absent_days'],
                    'nh_count': data['nh_count'],
                    'fh_count': data['fh_count'],
                } for data in attendance_data
            ]
            # Store additional session data for easy access
            request.session['dates'] = [date.strftime('%Y-%m-%d') for date in dates]
            request.session['start_date'] = start_date.strftime('%Y-%m-%d')
            request.session['end_date'] = end_date.strftime('%Y-%m-%d')
    else:
        form = MonthlyAttendanceForm()  # Handle the GET request by initializing an empty form

    # Render the attendance view template with context data
    context = {
        'form': form,
        'attendance_data': attendance_data,
        'dates': dates,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'audit_entries/monthly_attendance.html', context)




def render_to_pdf(template_src, context_dict):
    template = get_template(template_src)
    html = template.render(context_dict)
    result = BytesIO()
    pdf = pisa.CreatePDF(BytesIO(html.encode("UTF-8")), dest=result)
    if not pdf.err:
        return result.getvalue()
    return None

def download_pdf(request):
    attendance_data = request.session.get('attendance_data', [])
    dates = request.session.get('dates', [])
    start_date = request.session.get('start_date', '')
    end_date = request.session.get('end_date', '')

    context = {
        'attendance_data': attendance_data,
        'dates': dates,
        'start_date': start_date,
        'end_date': end_date,
    }

    pdf = render_to_pdf('audit_entries/monthly_attendance_pdf.html', context)
    if pdf:
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="monthly_attendance_report.pdf"'
        return response
    return HttpResponse("Error generating PDF", status=400)


import pandas as pd


def download_excel(request):
    attendance_data = request.session.get('attendance_data', [])
    dates = request.session.get('dates', [])
    
    data = []
    for data_entry in attendance_data:
        row = [
            data_entry['employee_code'],
            data_entry['employee_name'],
            data_entry['category'],
            data_entry['department']
        ]
        row.extend(data_entry['attendance'])
        data.append(row)
    
    columns = ['Employee Code', 'Employee Name', 'Category', 'Department'] + dates
    
    df = pd.DataFrame(data, columns=columns)
    
    with BytesIO() as buffer:
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Monthly Attendance')
        buffer.seek(0)
        return HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')



#time card



def render_to_pdf(template_src, context_dict):
    template = get_template(template_src)
    html = template.render(context_dict)
    result = BytesIO()
    pdf = pisa.CreatePDF(BytesIO(html.encode("UTF-8")), dest=result)
    if not pdf.err:
        return result.getvalue()
    return None


from django.shortcuts import render
from django.http import HttpResponse
from datetime import datetime, timedelta
import random
from django.shortcuts import render
from datetime import datetime, timedelta
from .forms import TimeCardForm
from .models import AuditAttendance

from django.shortcuts import render
from django.http import HttpResponse
from datetime import datetime, timedelta
import random
from django.shortcuts import render
from datetime import datetime, timedelta
from .forms import TimeCardForm
from .models import AuditAttendance

from datetime import datetime

from django.shortcuts import render
from datetime import timedelta
from .models import AuditAttendance, Employee, Holiday
from .forms import TimeCardForm
from datetime import timedelta
from django.shortcuts import render
from .models import AuditAttendance, Employee, Holiday
from .forms import TimeCardForm

def audit_timecard_view(request):
    form = TimeCardForm(request.POST or None)
    attendance_records = {}
    days_in_month = []
    month = None

    if form.is_valid():
        employee_code = form.cleaned_data.get('employee_code')
        category = form.cleaned_data.get('category')
        department = form.cleaned_data.get('department')
        month = form.cleaned_data.get('month')

        if month:
            # Calculate the start and end dates for the month
            first_day_of_month = month.replace(day=1)
            last_day_of_month = (month.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
            days_in_month = [first_day_of_month + timedelta(days=i) for i in range((last_day_of_month - first_day_of_month).days + 1)]

            # Build filters for querying attendance data
            filters = {'date__range': [first_day_of_month, last_day_of_month]}
            if employee_code:
                filters['employee__employee_code'] = employee_code.employee_code
            if category:
                filters['employee__department__category'] = category
            if department:
                filters['employee__department'] = department

            # Fetch attendance data for the filtered employees
            attendance_data = AuditAttendance.objects.filter(**filters).order_by('date')
            
            # Adjust employee query to include only those with pf_applicable = 'yes'
            employees = Employee.objects.filter(status='active', pf_esi_applicable='yes')

            # Apply employee filters based on the form inputs
            if employee_code:
                employees = employees.filter(employee_code=employee_code.employee_code)
            if category:
                employees = employees.filter(department__category=category)
            if department:
                employees = employees.filter(department=department)

            # Loop through each employee and build their timecard data
            for employee in employees:
                employee_days = []
                total_days = 0
                absent_days = 0
                total_hours = 0.0
                total_ot_hours = 0
                weekly_off_day = employee.weekly_off  # Get employee's weekly off day

                for date in days_in_month:
                    # Check if the day is a holiday
                    holiday = Holiday.objects.filter(date=date).first()  # Fetch holiday for the date
                    record = attendance_data.filter(employee=employee, date=date).first()
                    day_name = date.strftime('%A')  # Get the name of the day (e.g., Monday)

                    if holiday:
                        # If the date is a holiday (NH, FH, WH), show the holiday type
                        if holiday.holiday_type == 'FH':
                            holiday_display = f"Festival Holiday ({holiday.festival_name})"
                        else:
                            holiday_display = dict(Holiday.HOLIDAY_CHOICES).get(holiday.holiday_type)

                        employee_days.append({
                            'date': date.strftime('%Y-%m-%d'),  # Convert date to string
                            'intime': '',
                            'outtime': '',
                            'total_working_hours': 0,
                            'ot_hours': 0,
                            'shift': holiday.holiday_type,  # Show holiday type (WH, NH, FH)
                            'absent': holiday.holiday_type,  # Display the holiday type as 'absent'
                        })

                    elif record:
                        # If an attendance record exists for this date, process it
                        employee_days.append({
                            'date': date.strftime('%Y-%m-%d'),  # Convert date to string
                            'intime': record.intime.strftime('%H:%M:%S') if record.intime else '',
                            'outtime': record.outtime.strftime('%H:%M:%S') if record.outtime else '',
                            'total_working_hours': record.total_working_hours,
                            'ot_hours': record.ot_hours,
                            'shift': record.shift,
                            'absent': record.absent,
                        })

                        # Update counters based on attendance status
                        if record.absent == 'AB':
                            absent_days += 1
                        if record.absent == 'P':
                            total_days += 1
                        if record.total_working_hours:
                            total_hours += record.total_working_hours
                        if record.ot_hours:
                            total_ot_hours += record.ot_hours
                    

                    else:
                        # Handle days without attendance records (either weekly off or absent)
                        if day_name == weekly_off_day:
                            employee_days.append({
                                'date': date.strftime('%Y-%m-%d'),  # Convert date to string
                                'shift_time': '',
                                'intime': '',
                                'outtime': '',
                                'total_working_hours': 0,
                                'ot_hours': 0,
                                'shift': 'WO',
                                'absent': 'WO',
                            })
                        else:
                            employee_days.append({
                                'date': date.strftime('%Y-%m-%d'),  # Convert date to string
                                'shift_time': '',
                                'intime': '',
                                'outtime': '',
                                'total_working_hours': 0,
                                'ot_hours': 0,
                                'shift': '',
                                'absent': 'AB',
                            })
                            absent_days += 1

                # Store the attendance summary for each employee
                attendance_records[employee] = {
                    'days': list(enumerate(employee_days, start=1)),
                    'total_days': total_days,
                    'absent_days': absent_days,
                    'total_hours': total_hours,
                    'total_ot_hours': total_ot_hours,
                }

            # Save data to session for download or further processing
            request.session['attendance_data'] = [
                {
                    'employee_code': employee.employee_code,
                    'employee_name': employee.employee_name,
                    'category': employee.category.name if employee.category else '',
                    'department': employee.department.name if employee.department else '',
                    'attendance': data['days'],
                    'present_days': data['total_days'],
                    'absent_days': data['absent_days'],
                    'total_ot_hours': data['total_ot_hours'],
                } for employee, data in attendance_records.items()
            ]
            request.session['dates'] = [date.strftime('%Y-%m-%d') for date in days_in_month]  # Convert dates to strings
            request.session['month'] = month.strftime('%Y-%m')  # Convert month to string

    # Context to pass to the template
    context = {
        'form': form,
        'attendance_records': attendance_records,
        'days_in_month': days_in_month if month else [],
        'month': month,
    }

    return render(request, 'audit_entries/timecard.html', context)



def audit_salary_reports(request):
    return render(request, 'audit_salary/salary_reports.html')

from employee.models import Deduction, Incentive
from django.http import HttpResponse
from django.db.models import Sum
from employee.models import Employee, Wages, Incentive, Deduction

def audit_pf_salary_view(request):
    form = PfSalaryForm(request.POST or None)
    salary_details = []
    
    # Initialize total variables
    total_wages = 0
    total_deductions = 0
    total_net_salary = 0
    total_present_days = 0
    total_ot_hours = 0
    total_per_day_salary = 0
    total_paid_days = 0
    basic_salary = 0
    da_salary = 0
    hra_salary = 0

    if form.is_valid():
        selected_date = form.cleaned_data['month']
        year = selected_date.year
        month = selected_date.month

        # Calculate the first and last day of the month
        first_day_of_month = datetime(year, month, 1)
        last_day_of_month = (first_day_of_month.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
        days_in_month = last_day_of_month.day

        # Define working days based on the number of days in the month
        working_days = {
            31: 27,
            30: 26,
            29: 25,
            28: 24
        }.get(days_in_month, days_in_month)

        # Get the count of total holiday days in the month
        holiday_dates = list(Holiday.objects.filter(date__month=month, date__year=year).values_list('date', flat=True))
        total_holidays = len(holiday_dates)

        # Fetch attendance records
        attendance_records = AuditAttendance.objects.filter(
            date__range=[first_day_of_month, last_day_of_month],
            employee__pf_esi_applicable='yes',
            employee__status='active'
        )

        # Filter based on form data
        if form.cleaned_data.get('category'):
            attendance_records = attendance_records.filter(employee__category=form.cleaned_data['category'])
        if form.cleaned_data.get('department'):
            attendance_records = attendance_records.filter(employee__department=form.cleaned_data['department'])

        # Initialize employees dictionary
        employees = {}
        for record in attendance_records:
            if record.employee not in employees:
                employees[record.employee] = {
                    'attendance_records': [],
                    'total_p_days': 0,
                    'total_absent_days': 0,
                    'total_ot_hours': 0,
                    'deductions': {
                        'advance': 0.0,
                        'mess': 0.0,
                        'store': 0.0,
                        'others': 0.0,
                        'total': 0.0,
                    },
                    'holidays': total_holidays,
                    'working_days': working_days
                }

            ot_hours = record.amount / 8 if record.amount else 0
            employees[record.employee]['attendance_records'].append(record)

            # Update attendance counts
            if record.date in holiday_dates and record.absent == 'P':
                continue
            elif record.absent == 'P':
                employees[record.employee]['total_p_days'] += 1
            elif record.absent == 'AB':
                employees[record.employee]['total_absent_days'] += 1

            employees[record.employee]['total_ot_hours'] += ot_hours

        for i, (employee, data) in enumerate(employees.items(), start=1):
            try:
                wages = Wages.objects.get(employee=employee, start_date__lte=first_day_of_month, end_date__gte=last_day_of_month)
                per_day_salary = float(wages.per_day) if wages.per_day else 0.0
                basic_amount = float(wages.basic_amount) if wages.basic_amount else 0.0
                da_amount = float(wages.da_amount) if wages.da_amount else 0.0
                hra_amount = float(wages.hra_amount) if wages.hra_amount else 0.0
                other_allowance = float(wages.other_allowance) if wages.other_allowance else 0.0
            except Wages.DoesNotExist:
                per_day_salary = 0.0
                basic_amount = 0.0
                da_amount = 0.0
                hra_amount = 0.0
                other_allowance = 0.0

            total_days = data['total_p_days'] + data['holidays']
            capped_total_days = min(total_days, data['working_days'])
            total_wages = total_days * per_day_salary
            basic_salary = total_days * basic_amount 
            da_salary = total_days * da_amount 
            hra_salary = total_days * hra_amount 
            
            gross_amount=total_days * per_day_salary
            gross_amount = round(gross_amount)   # Round net salary to nearest 10

            epf_calculation = basic_amount + da_amount
            epf_wages = epf_calculation * capped_total_days
            epf_wages = min(epf_wages, 15000)
            pf_percentage= employee.pf_percentage
            pf_amount = round(epf_wages * pf_percentage / 100)
            pf_amount = round(pf_amount)   # Round net salary to nearest 10

            esi_wages = total_wages
            esi_wages = round(esi_wages)  # Round net salary to nearest 10
            esi_wages = min(esi_wages, 21000)
            esi_percentage =  employee.esi_percentage
            
            esi_amount = round(esi_wages * esi_percentage  / 100)
                

            esi_pf_total = esi_amount + pf_amount

            deductions = Deduction.objects.filter(
                employee=employee,
                date__month=month,
                date__year=year
            )
            for deduction in deductions:
                if deduction.deduction_type == 'advance':
                    employees[employee]['deductions']['advance'] += float(deduction.amount)
                elif deduction.deduction_type == 'mess':
                    employees[employee]['deductions']['mess'] += float(deduction.amount)
                elif deduction.deduction_type == 'store':
                    employees[employee]['deductions']['store'] += float(deduction.amount)
                elif deduction.deduction_type == 'others':
                    employees[employee]['deductions']['others'] += float(deduction.amount)

            employee_total_deduction = sum(employees[employee]['deductions'].values()) + esi_pf_total
            employees[employee]['deductions']['total'] = employee_total_deduction
            net_salary = gross_amount  - employee_total_deduction
             # Calculate total sums for all employees
            total_paid_days +=total_days
            total_wages += total_wages
            total_deductions += employee_total_deduction
            total_net_salary += net_salary
            total_present_days += data['total_p_days']
            total_ot_hours += data['total_ot_hours']
            total_per_day_salary += per_day_salary

            # Append salary details for rendering
            salary_details.append({
                'sl_no': i,
                'employee_code': employee.employee_code,
                'name': employee.employee_name,
                'pf':employee.pf_no ,
                'esi':employee.esi_no ,
                'department': employee.department.name,
                'designation': employee.designation.name,
                'pf_no': employee.pf_no,
                'esi_no': employee.esi_no,
                'total_present_days': data['total_p_days'],
                'total_days': total_days,
                'total_absent_days': data['total_absent_days'],
                'working_days': data['working_days'],
                'holidays': data['holidays'],
                'total_ot_hours': data['total_ot_hours'],
                'per_day_salary': per_day_salary,
                'total_wages': total_wages,
                'basic_amount': basic_amount,
                'da_amount': da_amount,
                'hra_amount': hra_amount,
                'other_allowance': other_allowance,
                'epf_wages': epf_wages,
                'pf_amount': pf_amount,
                'esi_wages': esi_wages,
                'esi_amount': esi_amount,
                'gross_amount': gross_amount,
                'total_deductions': employee_total_deduction,
                'net_salary': net_salary,
                'deductions': employees[employee]['deductions'],
            })
        salary_details = sorted(salary_details, key=lambda x: x['employee_code'])  # Sorting employee codes in ascending order 
           

        # Save data to session for download purposes
        request.session['salary_details'] = salary_details
        request.session['month'] = selected_date.strftime('%Y-%m')
        
        # Save total values in session
        request.session['total_wages'] = total_wages
        request.session['total_deductions'] = total_deductions
        request.session['total_net_salary'] = total_net_salary
        request.session['total_present_days'] = total_present_days
        request.session['total_paid_days'] = total_paid_days
        
        request.session['total_ot_hours'] = total_ot_hours
        request.session['total_per_day_salary'] = total_per_day_salary

        # Redirect to the details view
        return redirect(reverse('pf_salary_details_view'))

    return render(request, 'audit_salary/pf_salary.html', {
        'form': form,
        'salary_details': salary_details,
        'total_data': {
            'total_wages': total_wages,
            'total_deductions': total_deductions,
            'total_net_salary': total_net_salary,
            'total_present_days': total_present_days,
            'total_ot_hours': total_ot_hours,
            'total_per_day_salary': total_per_day_salary,
            'total_paid_days':total_paid_days
        }
    })

def pf_salary_details_view(request):
    # Retrieve salary details and selected month from the session
    salary_details = request.session.get('salary_details', [])
    selected_month = request.session.get('month', '')

    # Initialize total variables
    total_wages = 0
    total_present_days = 0
    total_paid_days=0
    total_gross_amount = 0
    total_esi_wages = 0
    total_epf_wages = 0
    total_esi_amount = 0
    total_pf_amount = 0
    total_deductions = {
        'advance': 0,
        'mess': 0,
        'store': 0,
        'others': 0,
    }
    total_net_salary = 0

    # Loop through salary details to calculate totals
    for detail in salary_details:
        total_wages += detail['total_wages']
        total_present_days += detail['total_present_days']
        total_gross_amount += detail['gross_amount']
        total_esi_wages += detail['esi_wages']
        total_epf_wages += detail['epf_wages']
        total_esi_amount += detail['esi_amount']
        total_pf_amount += detail['pf_amount']
        total_deductions['advance'] += detail['deductions']['advance']
        total_deductions['mess'] += detail['deductions']['mess']
        total_deductions['store'] += detail['deductions']['store']
        total_deductions['others'] += detail['deductions']['others']
        total_net_salary += detail['net_salary']

    # Render the response with salary details and totals
    return render(request, 'audit_salary/pf_salary_details_view.html', {
        'salary_details': salary_details,
        'selected_month': selected_month,
        'total_wages': total_wages,
        'total_present_days': total_present_days,
        'total_gross_amount': total_gross_amount,
        'total_esi_wages': total_esi_wages,
        'total_epf_wages': total_epf_wages,
        'total_esi_amount': total_esi_amount,
        'total_pf_amount': total_pf_amount,
        'total_deductions': total_deductions,
        'total_net_salary': total_net_salary,
        'total_paid_days':total_paid_days
    })


from django.http import HttpResponse
import pandas as pd
from io import BytesIO
from django.http import HttpResponse

def audit_pf_salary_excel(request):
    salary_details = request.session.get('salary_details', [])
    total_data = request.session.get('total_data', {})

    if not salary_details:
        return HttpResponse("No salary details to export.", status=400)

    # Prepare data for Excel export
    columns = [
        'Sl No', 'Employee Code', 'Name', 'Department', 'Designation',
        'Total Present Days','Holidays', 'Total Days',   'Total OT Hours', 
        'Per Day Salary', 'Total Wages', 'Basic Amount', 
        'DA Amount', 'HRA Amount', 'Other Allowance', 
        'EPF Wages', 'PF Amount', 'ESI Wages', 
        'ESI Amount', 'Gross Amount', 'Total Deductions', 
        'Net Salary'
    ]

    data = [
        [
            salary['sl_no'],
            salary['employee_code'],
            salary['name'],
            salary['department'],
            salary['designation'],
            salary['total_present_days'],
            salary['holidays'],
            salary['total_days'],
           
            
            salary['total_ot_hours'],
            salary['per_day_salary'],
            salary['total_wages'],
            salary['basic_amount'],
            salary['da_amount'],
            salary['hra_amount'],
            salary['other_allowance'],
            salary['epf_wages'],
            salary['pf_amount'],
            salary['esi_wages'],
            salary['esi_amount'],
            salary['gross_amount'],
            salary['total_deductions'],
            salary['net_salary']
        ] for salary in salary_details
    ]

    df = pd.DataFrame(data, columns=columns)

    # Create a summary row for totals
    total_row = {
        'Sl No': 'Total',
        'Employee Code': '',
        'Name': '',
        'Department': '',
        'Designation': '',
        'Total Present Days': df['Total Present Days'].sum(),
        'Holidays': df['Holidays'].sum(),
        'Total Days': df['Total Days'].sum(),
       
        
        'Total OT Hours': df['Total OT Hours'].sum(),
        'Per Day Salary': '',
        'Total Wages': df['Total Wages'].sum(),
        'Basic Amount': df['Basic Amount'].sum(),
        'DA Amount': df['DA Amount'].sum(),
        'HRA Amount': df['HRA Amount'].sum(),
        'Other Allowance': df['Other Allowance'].sum(),
        'EPF Wages': df['EPF Wages'].sum(),
        'PF Amount': df['PF Amount'].sum(),
        'ESI Wages': df['ESI Wages'].sum(),
        'ESI Amount': df['ESI Amount'].sum(),
        'Gross Amount': df['Gross Amount'].sum(),
        'Total Deductions': df['Total Deductions'].sum(),
        'Net Salary': df['Net Salary'].sum(),
    }

    # Create a DataFrame for the total row
    total_df = pd.DataFrame([total_row])

    # Concatenate the original DataFrame with the total DataFrame
    df = pd.concat([df, total_df], ignore_index=True)

    # Write to an Excel file
    with BytesIO() as buffer:
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='PF Salary Audit')
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="audit_pf_salary.xlsx"'
        return response




import pandas as pd
from io import BytesIO
from django.http import HttpResponse

def pf_ecr_form_statement_excel(request):
    # Retrieve salary details and totals from the session
    salary_details = request.session.get('salary_details', [])
    total_data = request.session.get('total_data', {})

    # Return an error if no salary details are found
    if not salary_details:
        return HttpResponse("No salary details to export.", status=400)

    # Prepare data for the DataFrame
    data = [
        [
            idx + 1,  # Use the index + 1 for 'Sl No'
            salary.get('pf_no', ''),
            salary.get('name', ''),
            salary.get('gross_amount', 0),
            salary.get('epf_wages', 0),
            salary.get('epf_wages', 0),
            salary.get('epf_wages', 0),
            salary.get('pf_amount', 0),
            salary.get('total_absent_days', 0),
            ''  # Signature field is left blank
        ] for idx, salary in enumerate(salary_details)
    ]

    # Create a DataFrame without explicit column names
    df = pd.DataFrame(data)

    # Write the DataFrame to an Excel file
    with BytesIO() as buffer:
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, header=False, sheet_name='PF ECR FORM')

        # Return the Excel file as a response
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="pf_ecr_form.xlsx"'
        return response













import pandas as pd
from io import BytesIO
from django.http import HttpResponse

def pf_esi_statement_excel(request):
    # Retrieve salary details and totals from the session
    salary_details = request.session.get('salary_details', [])
    total_data = request.session.get('total_data', {})

    # Return an error if no salary details are found
    if not salary_details:
        return HttpResponse("No salary details to export.", status=400)

    # Define the columns for the Excel file
    columns = [
        'Sl No', 'Employee Code', 'Name', 'PF No', 'ESI No',
        'Total Present Days', 'Total Absent Days','Gross Amount',
        'EPF Wages', 'ESI Wages', 'PF Amount', 'ESI Amount',
        'Net Salary', 'Signature'
    ]

    # Prepare data for the DataFrame using the index from enumerate as "Sl No"
    data = [
        [
            idx + 1,  # Use the index + 1 for 'Sl No'
            salary.get('employee_code', ''),
            salary.get('name', ''),
            format_number(salary.get('pf_no', '')),  # Format PF No to prevent .0
            format_number(salary.get('esi_no', '')),  # Format ESI No to prevent .0
            salary.get('total_days', 0),
            salary.get('total_absent_days', 0),
            salary.get('gross_amount', 0),
            salary.get('epf_wages', 0),
            salary.get('esi_wages', 0),
            salary.get('pf_amount', 0),
            salary.get('esi_amount', 0),
            salary.get('net_salary', 0),
            ''  # Signature field is left blank
        ] for idx, salary in enumerate(salary_details)
    ]

    # Create a DataFrame
    df = pd.DataFrame(data, columns=columns)

    # Create a summary row for totals
    total_row = {
        'Sl No': 'Total',
        'Employee Code': '',
        'Name': '',
        'PF No': '',
        'ESI No': '',
        'Total Present Days': df['Total Present Days'].sum(),
        'Total Absent Days': df['Total Absent Days'].sum(),
        'Gross Amount': df['Gross Amount'].sum(),
        'EPF Wages': df['EPF Wages'].sum(),
        'ESI Wages': df['ESI Wages'].sum(),
        'PF Amount': df['PF Amount'].sum(),
        'ESI Amount': df['ESI Amount'].sum(),
        'Net Salary': df['Net Salary'].sum(),
        'Signature': '',
    }

    # Append the total row to the DataFrame
    df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

    # Write the DataFrame to an Excel file
    with BytesIO() as buffer:
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='PF ESI STATEMENT')

        # Return the Excel file as a response
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="pf_esi_statement.xlsx"'
        return response

def format_number(number):
    """
    Function to format PF No and ESI No, removing the .0 part if present.
    Handles both float and string representations gracefully.
    """
    try:
        # If the number is a string ending with '.0'
        if isinstance(number, str) and number.endswith('.0'):
            number = number[:-2]  # Remove the '.0'
        # Convert float numbers to integer and then to string to remove decimal
        if isinstance(number, float):
            return str(int(number))
        return str(number)
    except (ValueError, TypeError):
        # If the conversion fails, return as an empty string or original value
        return str(number) if number else ''




from django.http import HttpResponse
from django.shortcuts import render
from django.http import HttpResponse
from django.shortcuts import render
def audit_pf_salary_pdf(request):
    # Retrieve salary details from the session
    selected_month = request.session.get('month', '')
    salary_details = request.session.get('salary_details', [])
    
    # Initialize total variables
    total_wages = 0
    total_present_days = 0
    total_gross_amount = 0
    total_esi_wages = 0
    total_epf_wages = 0
    total_esi_amount = 0
    total_pf_amount = 0
    total_holidays = 0
    total_paid_days = 0
    total_deductions = {
        'advance': 0,
        'mess': 0,
        'store': 0,
        'others': 0,
    }
    total_net_salary = 0

    # Loop through salary details to calculate totals
    for detail in salary_details:
        total_wages += detail['total_wages']
        total_present_days += detail['total_present_days']
        total_gross_amount += detail['gross_amount']
        total_esi_wages += detail['esi_wages']
        total_epf_wages += detail['epf_wages']
        total_esi_amount += detail['esi_amount']
        total_pf_amount += detail['pf_amount']
        total_holidays += detail.get('holidays', 0)
        total_deductions['advance'] += detail['deductions'].get('advance', 0)
        total_deductions['mess'] += detail['deductions'].get('mess', 0)
        total_deductions['store'] += detail['deductions'].get('store', 0)
        total_deductions['others'] += detail['deductions'].get('others', 0)
        total_net_salary += detail['net_salary']

    # Calculate total deductions
    total_deductions_sum = (
        sum(total_deductions.values()) + total_esi_amount + total_pf_amount
    )


    # Prepare total data dictionary
    total_data = {
        'total_wages': total_wages,
        'total_present_days': total_present_days,
        'total_gross_amount': total_gross_amount,
        'total_esi_wages': total_esi_wages,
        'total_epf_wages': total_epf_wages,
        'total_esi_amount': total_esi_amount,
        'total_pf_amount': total_pf_amount,
        'total_holidays': total_holidays,
        'total_deductions': total_deductions,
        'total_deductions_sum': total_deductions_sum,  # Add total deductions to context
        'total_net_salary': total_net_salary,
        'total_paid_days':total_paid_days,
       
    }

    context = {
        'salary_details': salary_details,
        'total_data': total_data,
         'month':selected_month,
    }

    pdf = render_to_pdf('audit_salary/pf_salary_pdf.html', context)
    if pdf:
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="audit_pf_salary.pdf"'
        return response
    
    return HttpResponse("Error generating PDF", status=400)
from datetime import datetime, timedelta
from decimal import Decimal

from django.shortcuts import render
from .forms import PfSalaryForm
from datetime import datetime, timedelta
from decimal import Decimal
from datetime import datetime, timedelta
from decimal import Decimal

from django.shortcuts import render
from .forms import PfSalaryForm

from django.shortcuts import render

from django.db.models import Sum
from decimal import Decimal
from datetime import datetime, timedelta

def bank_details_pf_salary(request):

    form = PfSalaryForm(request.POST or None)
    salary_details = []
    
    # Initialize total variables
    total_wages = 0
    total_deductions = 0
    total_net_salary = 0
    total_present_days = 0
    total_ot_hours = 0
    total_per_day_salary = 0

    if form.is_valid():
        selected_date = form.cleaned_data['month']
        year = selected_date.year
        month = selected_date.month

        # Calculate the first and last day of the month
        first_day_of_month = datetime(year, month, 1)
        last_day_of_month = (first_day_of_month.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
        days_in_month = last_day_of_month.day

        # Define working days based on the number of days in the month
        working_days = {
            31: 27,
            30: 26,
            29: 25,
            28: 24
        }.get(days_in_month, days_in_month)

        # Get the count of total holiday days in the month
        holiday_dates = list(Holiday.objects.filter(date__month=month, date__year=year).values_list('date', flat=True))
        total_holidays = len(holiday_dates)

        # Fetch attendance records
        attendance_records = AuditAttendance.objects.filter(
            date__range=[first_day_of_month, last_day_of_month],
            employee__pf_esi_applicable='yes',
            employee__status='active'
        )

        # Filter based on form data
        if form.cleaned_data.get('category'):
            attendance_records = attendance_records.filter(employee__category=form.cleaned_data['category'])
        if form.cleaned_data.get('department'):
            attendance_records = attendance_records.filter(employee__department=form.cleaned_data['department'])

        # Initialize employees dictionary
        employees = {}
        for record in attendance_records:
            if record.employee not in employees:
                employees[record.employee] = {
                    'attendance_records': [],
                    'total_p_days': 0,
                    'total_absent_days': 0,
                    'total_ot_hours': 0,
                    'deductions': {
                        'advance': 0.0,
                        'mess': 0.0,
                        'store': 0.0,
                        'others': 0.0,
                        'total': 0.0,
                    },
                    'holidays': total_holidays,
                    'working_days': working_days
                }

            ot_hours = record.amount / 8 if record.amount else 0
            employees[record.employee]['attendance_records'].append(record)

            # Update attendance counts
            if record.date in holiday_dates and record.absent == 'P':
                continue
            elif record.absent == 'P':
                employees[record.employee]['total_p_days'] += 1
            elif record.absent == 'AB':
                employees[record.employee]['total_absent_days'] += 1

            employees[record.employee]['total_ot_hours'] += ot_hours

        for i, (employee, data) in enumerate(employees.items(), start=1):
            try:
                wages = Wages.objects.get(employee=employee, start_date__lte=first_day_of_month, end_date__gte=last_day_of_month)
                per_day_salary = float(wages.per_day) if wages.per_day else 0.0
                basic_amount = float(wages.basic_amount) if wages.basic_amount else 0.0
                da_amount = float(wages.da_amount) if wages.da_amount else 0.0
                hra_amount = float(wages.hra_amount) if wages.hra_amount else 0.0
                other_allowance = float(wages.other_allowance) if wages.other_allowance else 0.0
            except Wages.DoesNotExist:
                per_day_salary = 0.0
                basic_amount = 0.0
                da_amount = 0.0
                hra_amount = 0.0
                other_allowance = 0.0

            total_days = data['total_p_days'] + data['holidays']
            capped_total_days = min(total_days, data['working_days'])
            total_wages = total_days * per_day_salary
            ot_amount = data['total_ot_hours'] * per_day_salary
            gross_amount=total_days * per_day_salary
            gross_amount = round(gross_amount)   # Round net salary to nearest 10

            epf_calculation = basic_amount + da_amount
            epf_wages = epf_calculation * capped_total_days
            epf_wages = min(epf_wages, 15000)
            pf_percentage= employee.pf_percentage
            pf_amount = round(epf_wages * pf_percentage / 100)
            pf_amount = round(pf_amount)   # Round net salary to nearest 10

            esi_wages = total_wages
            esi_wages = round(esi_wages)   # Round net salary to nearest 10
            esi_wages = min(esi_wages, 21000)
            esi_percentage =  employee.esi_percentage
            
            esi_amount = round(esi_wages * esi_percentage  / 100)    

            esi_pf_total = esi_amount + pf_amount

            deductions = Deduction.objects.filter(
                employee=employee,
                date__month=month,
                date__year=year
            )
            for deduction in deductions:
                if deduction.deduction_type == 'advance':
                    employees[employee]['deductions']['advance'] += float(deduction.amount)
                elif deduction.deduction_type == 'mess':
                    employees[employee]['deductions']['mess'] += float(deduction.amount)
                elif deduction.deduction_type == 'store':
                    employees[employee]['deductions']['store'] += float(deduction.amount)
                elif deduction.deduction_type == 'others':
                    employees[employee]['deductions']['others'] += float(deduction.amount)

            employee_total_deduction = sum(employees[employee]['deductions'].values()) + esi_pf_total
            employees[employee]['deductions']['total'] = employee_total_deduction
            net_salary = gross_amount + ot_amount - employee_total_deduction
             # Calculate total sums for all employees
            total_wages += total_wages
            total_deductions += employee_total_deduction
            total_net_salary += net_salary
            total_present_days += data['total_p_days']
            total_ot_hours += data['total_ot_hours']
            total_per_day_salary += per_day_salary

            
        
            



            # Append salary details for the employee
            salary_details.append({
                'sl_no': i,
                'employee_code': employee.employee_code,
                'name': employee.employee_name,
                'department': employee.department.name,
               
                'bank':employee.bank_name, 
                'branch':employee.branch,
                'account_no':employee.account_no, 
                'account_holder': employee.account_holder, 

                'ifsc_code':employee.ifsc_code,
                'total_present_days': data['total_p_days'],
                'total_days': total_days,
                'total_absent_days': data['total_absent_days'],
                'working_days': data['working_days'],
                'holidays': data['holidays'],
                'total_ot_hours': data['total_ot_hours'],
                'per_day_salary': per_day_salary,
                'total_wages': total_wages,
                'basic_amount': per_day_salary,
                'da_amount': 0,  # DA amount if needed
                'hra_amount': hra_amount,
                'other_allowance': other_allowance,
                'epf_wages': epf_wages,
                'pf_amount': pf_amount,
                'esi_wages': esi_wages,
                'esi_amount': esi_amount,
                'gross_amount': total_wages + ot_amount,
                'total_deductions': total_deductions,
                'net_salary': net_salary,
                'deductions': employees[employee]['deductions'],
                
            })
        salary_details = sorted(salary_details, key=lambda x: x['employee_code'])  # Sorting employee codes in ascending order 
         # Store in session
        request.session['salary_details'] = salary_details
        request.session['total_net_salary'] = total_net_salary

     

    return render(request, 'audit_salary/bank_details_pf_salary_form.html', {'salary_details': salary_details, 'form': form, 'total_net_salary':total_net_salary,})
import pandas as pd
from io import BytesIO
from django.http import HttpResponse

def bank_audit_pf_salary_excel(request):
    # Retrieve salary details from the session
    salary_details = request.session.get('salary_details', [])
    
    if not salary_details:
        return HttpResponse("No salary details to export.", status=400)

    # Prepare data for Excel export
    columns = [
        'Sl No', 'Employee Code', 'Name','Account Holder' , 'Account Number','Bank', 'Branch',
        'IFSC', 'Net Salary'
    ]

    # Extract the relevant data for export and ensure account numbers are formatted correctly
    data = [
        [
            salary['sl_no'],
            salary['employee_code'],
            salary['name'],
            salary['account_holder'],
            format_account_number(salary['account_no']),
            salary['bank'],
            salary['branch'],
      
            salary['ifsc_code'],
            salary['net_salary']
        ] for salary in salary_details
    ]

    # Create a DataFrame
    df = pd.DataFrame(data, columns=columns)

    # Create a summary row for totals
    total_row = {
        'Sl No': 'Total',
        'Employee Code': '',
        'Name': '',
        'Account Holder':'',
         'Account Number': '',  # Leave empty for total row
        'Bank': '',
        'Branch': '',
       
        'IFSC': '',
        'Net Salary': df['Net Salary'].sum()
    }

    # Create a DataFrame for the total row
    total_df = pd.DataFrame([total_row])

    # Concatenate the original DataFrame with the total DataFrame
    df = pd.concat([df, total_df], ignore_index=True)

    # Write to an Excel file
    with BytesIO() as buffer:
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Bank Salary Details')
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="bank_salary_details.xlsx"'
        return response

def format_account_number(account_number):
    """Function to format account number, removing the .0 part."""
    try:
        # Check if the account number is a float represented as a string (e.g., '1131178000015834.0')
        if isinstance(account_number, str) and account_number.endswith('.0'):
            account_number = account_number[:-2]  # Remove the '.0'
        # Convert float account number to an integer and then to string (handles floats like 1131178000015834.0)
        return str(int(float(account_number)))
    except (ValueError, TypeError):
        # If the conversion fails, return as an empty string or original value
        return str(account_number) if account_number else ''

from datetime import datetime, timedelta
from django.shortcuts import render
from .forms import EmployeeSalaryForm
from django.shortcuts import render
from django.http import HttpResponse
from django.template.loader import get_template
from io import BytesIO
from xhtml2pdf import pisa
from datetime import datetime, timedelta
import decimal
from django.shortcuts import render
from django.http import HttpResponse
from django.template.loader import get_template
from io import BytesIO
from xhtml2pdf import pisa
from datetime import datetime, timedelta
from io import BytesIO
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
from .forms import EmployeeSalaryForm

from xhtml2pdf import pisa
from django.template.loader import get_template
from decimal import Decimal
from datetime import datetime, timedelta


from num2words import num2words 

from num2words import num2words 
from datetime import datetime, timedelta
from decimal import Decimal
from django.shortcuts import render
from num2words import num2words
from django.shortcuts import render
from datetime import datetime, timedelta
from decimal import Decimal
from datetime import datetime, timedelta
from num2words import num2words  # Import the num2words library for converting numbers to words

def employee_salary_view(request):
    form = EmployeeSalaryForm(request.POST or None)
    salary_details = []

    # Initialize total variables
    total_wages = 0
    total_deductions = 0
    total_net_salary = 0
    total_present_days = 0
    total_ot_hours = 0
    total_per_day_salary = 0
    total_paid_days = 0

    if form.is_valid():
        selected_date = form.cleaned_data['month']
        selected_employee = form.cleaned_data.get('employee_code', None)  # This can be 'None' for All Employees
        category = form.cleaned_data.get('category', None)

        year = selected_date.year
        month = selected_date.month

        # Calculate the first and last day of the month
        first_day_of_month = datetime(year, month, 1)
        last_day_of_month = (first_day_of_month.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
        days_in_month = last_day_of_month.day

        # Define working days based on the number of days in the month
        working_days = {
            31: 27,
            30: 26,
            29: 25,
            28: 24
        }.get(days_in_month, days_in_month)

        # Get the count of total holiday days in the month
        holiday_dates = list(Holiday.objects.filter(date__month=month, date__year=year).values_list('date', flat=True))
        total_holidays = len(holiday_dates)

        # Fetch attendance records for the selected employee or all employees
        attendance_records = AuditAttendance.objects.filter(
            date__range=[first_day_of_month, last_day_of_month],
            employee__pf_esi_applicable='yes',
            employee__status='active'
        )

        # If a specific employee is selected, filter for that employee
        if selected_employee:
            attendance_records = attendance_records.filter(employee=selected_employee)
        # Otherwise, proceed with all employees
        if category:
            attendance_records = attendance_records.filter(employee__category=category)

        # Initialize employees dictionary
        employees = {}
        for record in attendance_records:
            if record.employee not in employees:
                employees[record.employee] = {
                    'attendance_records': [],
                    'total_p_days': 0,
                    'total_absent_days': 0,
                    'total_ot_hours': 0,
                    'deductions': {
                        'advance': 0.0,
                        'mess': 0.0,
                        'store': 0.0,
                        'others': 0.0,
                        'total': 0.0,
                    },
                    'holidays': total_holidays,
                    'working_days': working_days
                }

            ot_hours = record.amount / 8 if record.amount else 0
            employees[record.employee]['attendance_records'].append(record)

            # Update attendance counts
            if record.date in holiday_dates and record.absent == 'P':
                continue
            elif record.absent == 'P':
                employees[record.employee]['total_p_days'] += 1
            elif record.absent == 'AB':
                employees[record.employee]['total_absent_days'] += 1

            employees[record.employee]['total_ot_hours'] += ot_hours

        # Process each employee's salary
        for i, (employee, data) in enumerate(employees.items(), start=1):
            try:
                wages = Wages.objects.get(employee=employee, start_date__lte=first_day_of_month, end_date__gte=last_day_of_month)
                per_day_salary = float(wages.per_day) if wages.per_day else 0.0
                basic_amount = float(wages.basic_amount) if wages.basic_amount else 0.0
                da_amount = float(wages.da_amount) if wages.da_amount else 0.0
                hra_amount = float(wages.hra_amount) if wages.hra_amount else 0.0
                other_allowance = float(wages.other_allowance) if wages.other_allowance else 0.0
            except Wages.DoesNotExist:
                per_day_salary = 0.0
                basic_amount = 0.0
                da_amount = 0.0
                hra_amount = 0.0
                other_allowance = 0.0

            total_days = data['total_p_days'] + data['holidays']
            capped_total_days = min(total_days, data['working_days'])
            total_wages = total_days * per_day_salary
            ot_amount = data['total_ot_hours'] * per_day_salary
            gross_amount = round(total_wages / 10) * 10  # Round gross salary to nearest 10

            # EPF and ESI calculations
            epf_calculation = basic_amount + da_amount
            epf_wages = epf_calculation * capped_total_days
            epf_wages = min(epf_wages, 15000)
            pf_percentage= employee.pf_percentage
            pf_amount = round(epf_wages * pf_percentage / 100)
            pf_amount = round(pf_amount / 10) * 10  # Round net salary to nearest 10

            esi_wages = total_wages
            esi_wages = round(esi_wages / 10) * 10  # Round net salary to nearest 10
            esi_wages = min(esi_wages, 21000)
            esi_percentage =  employee.esi_percentage
            
            esi_amount = round(esi_wages * esi_percentage  / 100)

            esi_pf_total = esi_amount + pf_amount

            # Fetch deductions
            deductions = Deduction.objects.filter(
                employee=employee,
                date__month=month,
                date__year=year
            )
            for deduction in deductions:
                if deduction.deduction_type == 'advance':
                    employees[employee]['deductions']['advance'] += float(deduction.amount)
                elif deduction.deduction_type == 'mess':
                    employees[employee]['deductions']['mess'] += float(deduction.amount)
                elif deduction.deduction_type == 'store':
                    employees[employee]['deductions']['store'] += float(deduction.amount)
                elif deduction.deduction_type == 'others':
                    employees[employee]['deductions']['others'] += float(deduction.amount)

            employee_total_deduction = sum(employees[employee]['deductions'].values()) + esi_pf_total
            employees[employee]['deductions']['total'] = employee_total_deduction
            net_salary = gross_amount + ot_amount - employee_total_deduction

            # Format the dates for display
            from_date = first_day_of_month.strftime('%d-%m-%Y')
            to_date = last_day_of_month.strftime('%d-%m-%Y')

            # Calculate Date of Payment
            payment_date = datetime(year, month, 7)
            if payment_date.weekday() == 6:
                payment_date -= timedelta(days=1)
            payment_date_str = payment_date.strftime('%d-%m-%Y')

            # Convert net salary to words
            net_salary_in_words = num2words(net_salary, lang='en').capitalize() + " only"

            # Aggregate totals
            total_paid_days += total_days
            total_wages += total_wages
            total_deductions += employee_total_deduction
            total_net_salary += net_salary
            total_present_days += data['total_p_days']
            total_ot_hours += data['total_ot_hours']
            total_per_day_salary += per_day_salary

            # Append salary details for rendering
            salary_details.append({
                'sl_no': i,
                'employee_code': employee.employee_code,
                'name': employee.employee_name,
                'department': employee.department.name,
                'employee_name': employee.employee_name,
                'father_or_husband_name':employee.father_or_husband_name,
                'designation': employee.designation.name,
                'pf_number': employee.pf_no,
                'esi_no': employee.esi_no,
                'account_no': employee.account_no,
                'total_present_days': data['total_p_days'],
                'total_days': total_days,
                'total_absent_days': data['total_absent_days'],
                'working_days': data['working_days'],
                'holidays': data['holidays'],
                'total_ot_hours': data['total_ot_hours'],
                'per_day_salary': per_day_salary,
                'total_wages': total_wages,
                'basic_amount': basic_amount,
                'da_amount': da_amount,
                'hra_amount': hra_amount,
                'other_allowance': other_allowance,
                'epf_wages': epf_wages,
                'pf_amount': pf_amount,
                'esi_wages': esi_wages,
                'esi_amount': esi_amount,
                'gross_amount': gross_amount,
                'total_deductions': employee_total_deduction,
                'net_salary': net_salary,
                'net_salary_in_words': net_salary_in_words,
                'deductions': employees[employee]['deductions'],
                'from_date': from_date,
                'to_date': to_date,
                'payment_date': payment_date_str,
            })

    return render(request, 'audit_salary/employee_salary.html', {
        'form': form,
        'salary_details': salary_details,
    })


def render_to_pdf(template_src, context_dict):
    template = get_template(template_src)
    html = template.render(context_dict)
    result = BytesIO()
    pdf = pisa.CreatePDF(BytesIO(html.encode("UTF-8")), dest=result)
    if not pdf.err:
        return result.getvalue()
    return None

def employee_salary_download_pdf(request):
    salary_details = request.session.get('salary_details', [])
    month = request.session.get('month', '')

    if not salary_details:
        return HttpResponse("No salary details to export.", status=400)

    context = {
        'salary_details': salary_details,
        'month': month,
    }

    pdf = render_to_pdf('audit_salary/payslip.html', context)
    if pdf:
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="employee_salary_{month}.pdf"'
        return response
    return HttpResponse("Error generating PDF", status=400)




###updation upload employee attemdance#######
import pandas as pd
from django.shortcuts import render, redirect
from django.contrib import messages
import pandas as pd
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import AuditAttendance

def upload_attendance_employee_wise(request):
    if request.method == 'POST':
        attendance_file = request.FILES.get('attendance_file')

        if attendance_file:
            try:
                # Read the Excel file
                df = pd.read_excel(attendance_file)

                # Track whether any valid records were processed
                records_processed = False

                for _, row in df.iterrows():
                    employee_code = row.get('Employee Code')
                    date = row.get('Date')
                    absent = row.get('Absent')

                    # Skip the row if employee_code is missing
                    if not employee_code:
                        messages.warning(request, "Employee Code is missing in a row. This row will be skipped.")
                        continue  # Skip this row

                    # Check if the employee exists
                    try:
                        employee = Employee.objects.get(employee_code=employee_code)
                    except Employee.DoesNotExist:
                        messages.warning(request, f"Employee code {employee_code} not found. Record skipped.")
                        continue  # Skip this row

                    # Update or create the attendance record
                    AuditAttendance.objects.update_or_create(
                        employee=employee,
                        date=date,
                        defaults={'absent': absent}
                    )

                    records_processed = True

                if records_processed:
                    messages.success(request, "PF Employee Attendance records processed successfully.")
                else:
                    messages.info(request, "No valid records were uploaded.")

            except Exception as e:
                messages.error(request, f"Error processing the file: {e}")

            return redirect('pf_upload_attendance_employee_wise')  # Redirect to the upload page

    return render(request, 'audit_entries/upload_attendance_employee_wise.html')  # Render the upload template

###################################################






# Helper function to convert Decimal to float
def convert_decimals_to_float(data):
    if isinstance(data, dict):
        return {key: convert_decimals_to_float(value) for key, value in data.items()}
    elif isinstance(data, list):
        return [convert_decimals_to_float(item) for item in data]
    elif isinstance(data, Decimal):
        return float(data)
    return data


from .forms import MonthForm

def pf_category_totals_view(request):
    form = MonthForm(request.POST or None)
    salary_details = []
    selected_month_name = None
    
    # Initialize overall totals
    total_paid_days = 0
    total_ot_hours = 0
    all_total_wages = 0
    
    total_gross_amount = 0
    all_total_deductions = 0
    total_incentives = 0
    total_advance_deductions = 0
    total_canteen_deductions = 0
    total_mess_deductions = 0
    total_store_deductions = 0
    total_other_deductions = 0
    total_net_salary = 0
    total_ot_amount = 0
    total_wages = 0
    total_deductions = 0
    total_net_salary = 0
    total_present_days = 0
   
    total_per_day_salary = 0
    total_paid_days = 0
    total_pf_amount = 0
    basic_salary = 0
    da_salary = 0
    hra_salary = 0
     

    # Initialize category and department totals
    category_totals = {}
    department_totals = {}

    # Overall totals for the organization
    overall_totals = {
        'total_employees': 0,
        'total_advance_deductions': 0,
        'total_mess_deductions': 0,
        'total_store_deductions': 0,
        'total_other_deductions': 0,
        'total_deductions': 0,
        'net_salary_total': 0,
        'total_incentives': 0,
        'total_gross_amount': 0,
        'total_days':0,
        'all_employees':0,
        'total_per_day_salary':0,
        'total_shift_wages':0,
        'total_basic_amount':0,
        'total_da_amount':0,
        'total_hra_amount':0,
        'total_pf_amount':0,
        'total_esi_amount': 0 ,

       
       
        

    }

    # Initialize the total employees count for the entire organization
    overall_total_employees = 0
    selected_month = None
    selected_year = None
    month = None
    year = None
    month_name = None

    if form.is_valid():
        selected_date = form.cleaned_data['month']
        year = selected_date.year
        month = selected_date.month
        
        selected_year = selected_date.year
        selected_month = selected_date.month
        month_name = selected_date.strftime('%B')

        # Calculate the first and last day of the month
        first_day_of_month = datetime(year, month, 1)
        last_day_of_month = (first_day_of_month.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
        days_in_month = last_day_of_month.day

        # Define working days based on the number of days in the month
        working_days = {
            31: 27,
            30: 26,
            29: 25,
            28: 24
        }.get(days_in_month, days_in_month)

        # Get the count of total holiday days in the month
        holiday_dates = list(Holiday.objects.filter(date__month=month, date__year=year).values_list('date', flat=True))
        total_holidays = len(holiday_dates)

        # Fetch attendance records
        attendance_records = AuditAttendance.objects.filter(
            date__range=[first_day_of_month, last_day_of_month],
            employee__pf_esi_applicable='yes',
            employee__status='active'
        )

        # Filter based on form data
        if form.cleaned_data.get('category'):
            attendance_records = attendance_records.filter(employee__category=form.cleaned_data['category'])
        if form.cleaned_data.get('department'):
            attendance_records = attendance_records.filter(employee__department=form.cleaned_data['department'])

        # Initialize employees dictionary
        employees = {}
        for record in attendance_records:
            if record.employee not in employees:
                employees[record.employee] = {
                    'attendance_records': [],
                    'total_p_days': 0,
                    'total_absent_days': 0,
                    'total_ot_hours': 0,
                    'deductions': {
                        'advance': 0.0,
                        'mess': 0.0,
                        'store': 0.0,
                        'others': 0.0,
                        'total': 0.0,
                    },
                    'holidays': total_holidays,
                    'working_days': working_days
                }

            ot_hours = record.amount / 8 if record.amount else 0
            employees[record.employee]['attendance_records'].append(record)

            # Update attendance counts
            if record.date in holiday_dates and record.absent == 'P':
                continue
            elif record.absent == 'P':
                employees[record.employee]['total_p_days'] += 1
            elif record.absent == 'AB':
                employees[record.employee]['total_absent_days'] += 1

            employees[record.employee]['total_ot_hours'] += ot_hours

        for i, (employee, data) in enumerate(employees.items(), start=1):
            try:
                wages = Wages.objects.get(employee=employee, start_date__lte=first_day_of_month, end_date__gte=last_day_of_month)
                per_day_salary = float(wages.per_day) if wages.per_day else 0.0
                basic_amount = float(wages.basic_amount) if wages.basic_amount else 0.0
                da_amount = float(wages.da_amount) if wages.da_amount else 0.0
                hra_amount = float(wages.hra_amount) if wages.hra_amount else 0.0
                other_allowance = float(wages.other_allowance) if wages.other_allowance else 0.0
                rent = float(wages.rent) if wages.rent else 0.0
            except Wages.DoesNotExist:
                per_day_salary = 0.0
                basic_amount = 0.0
                da_amount = 0.0
                hra_amount = 0.0
                other_allowance = 0.0

            overall_total_employees += 1

            total_days = data['total_p_days'] + data['holidays']
            capped_total_days = min(total_days, data['working_days'])
            total_wages = total_days * per_day_salary
             
            basic_salary = total_days * basic_amount 
            da_salary = total_days * da_amount 
            hra_salary = total_days * hra_amount 
           
            gross_amount=total_days * per_day_salary
            gross_amount = round(gross_amount)   # Round net salary to nearest 10

            epf_calculation = basic_amount + da_amount
            epf_wages = epf_calculation * capped_total_days
            epf_wages = min(epf_wages, 15000)
            pf_percentage= employee.pf_percentage
            pf_amount = round(epf_wages * pf_percentage / 100)
            

            esi_wages = total_wages
            esi_wages = round(esi_wages)  # Round net salary to nearest 10
            esi_wages = min(esi_wages, 21000)
            esi_percentage =  employee.esi_percentage
            
            esi_amount = round(esi_wages * esi_percentage  / 100)
                

            esi_pf_total = esi_amount + pf_amount

            deductions = Deduction.objects.filter(
                employee=employee,
                date__month=month,
                date__year=year
            )
            for deduction in deductions:
                if deduction.deduction_type == 'advance':
                    employees[employee]['deductions']['advance'] += float(deduction.amount)
                elif deduction.deduction_type == 'mess':
                    employees[employee]['deductions']['mess'] += float(deduction.amount)
                elif deduction.deduction_type == 'store':
                    employees[employee]['deductions']['store'] += float(deduction.amount)
                elif deduction.deduction_type == 'others':
                    employees[employee]['deductions']['others'] += float(deduction.amount)

            employee_total_deduction = sum(employees[employee]['deductions'].values()) + esi_pf_total
            employees[employee]['deductions']['total'] = employee_total_deduction
            net_salary = gross_amount  - employee_total_deduction

            # Store salary details in a list
            salary_details.append({
                'sl_no': i,
                'employee_code': employee.employee_code,
                'name': employee.employee_name,
                'department': employee.department.name,
                'designation': employee.designation.name if employee.designation else '',
                'hra_amount': float(hra_amount),
                'other_allowance': float(other_allowance),
                'rent': float(rent),
                'total_p_days': float(data['total_p_days']),
                'total_ot_hours': float(data['total_ot_hours']),
                'per_day_salary': float(per_day_salary),
                'total_wages': float(total_wages),
               
                'gross_amount': float(gross_amount),
                'total_deductions': float(total_deductions),
               
                'net_salary': float(net_salary),
          
                'deductions': {key: float(value) for key, value in data['deductions'].items()},
            })

            # Update overall totals
            total_paid_days += data['total_p_days']
            total_ot_hours += data['total_ot_hours']
            all_total_wages += total_wages
            
            total_gross_amount += gross_amount
            all_total_deductions += total_deductions
            
            
            total_net_salary += net_salary  # Accumulate net salary here
       
            
        
            # Accumulate department totals (new logic)
            department_name = employee.department.name if employee.department else 'Uncategorized'
            if department_name not in department_totals:
                department_totals[department_name] = {
                    'total_per_day_salary': 0,  # Track total per day salary
                    'total_employees': 0,  # Track the total number of employees
                    'shift_wages':0, 
             
                    'advance_total': 0,  # Track total advance deductions
                    
                    'mess_total': 0,  # Track total mess deductions
                    'store_total': 0,  # Track total store deductions
                    'others_total': 0,  # Track total other deductions
                    'total_deductions': 0,  # Track total deductions in the department
                    'net_salary_total': 0,
                    
                    'total_gross_amount': 0,
                    'total_paid_days': 0,  # Ensure this key is initialized
                    'total_rent': 0, 
                    'total_ot_hours':0,
                    'total_basic_amount':0,
        'total_da_amount':0,
        'total_hra_amount':0,
        'total_pf_amount':0,
        'total_esi_amount': 0 ,
                }
            
            # Update department totals for this employee
            department_totals[department_name]['total_pf_amount'] += pf_amount
            department_totals[department_name]['total_esi_amount'] += esi_amount
            department_totals[department_name]['total_ot_hours'] += data['total_ot_hours']
            department_totals[department_name]['total_basic_amount']  += basic_salary
            department_totals[department_name]['total_da_amount']  += da_salary
            department_totals[department_name]['total_hra_amount']  += hra_salary
        
            department_totals[department_name]['shift_wages'] += total_wages
            department_totals[department_name]['total_per_day_salary'] += per_day_salary
            
            department_totals[department_name]['total_paid_days'] += data['total_p_days']
            department_totals[department_name]['total_employees'] += 1
            department_totals[department_name]['advance_total'] += data['deductions']['advance']
          
            department_totals[department_name]['mess_total'] += data['deductions']['mess']
            department_totals[department_name]['store_total'] += data['deductions']['store']
            department_totals[department_name]['others_total'] += data['deductions']['others']
            department_totals[department_name]['total_deductions'] += total_deductions
            department_totals[department_name]['net_salary_total'] += net_salary
         
            department_totals[department_name]['total_gross_amount'] += gross_amount
 

            # Update overall totals for the entire organization
        
            overall_totals['total_pf_amount'] += pf_amount
            overall_totals['total_esi_amount'] += esi_amount
            overall_totals['total_basic_amount']  += basic_salary
            overall_totals['total_da_amount']  += da_salary
            overall_totals['total_hra_amount']  += hra_salary
            overall_totals['total_shift_wages']  += total_wages
            overall_totals['total_days']  += data['total_p_days']
            overall_totals['total_per_day_salary'] += per_day_salary
            overall_totals['total_advance_deductions'] += data['deductions']['advance']
            overall_totals['total_mess_deductions'] += data['deductions']['mess']
            overall_totals['total_store_deductions'] += data['deductions']['store']
            overall_totals['total_other_deductions'] += data['deductions']['others']
            overall_totals['total_deductions'] += total_deductions
            overall_totals['net_salary_total'] += net_salary
        
            overall_totals['total_gross_amount'] += gross_amount

        # Update the overall employee count
            overall_totals['total_employees'] = overall_total_employees

    department_totals = dict(sorted(department_totals.items()))

    request.session['department_totals'] = convert_decimals_to_float(department_totals)
    request.session['overall_totals'] = convert_decimals_to_float(overall_totals)
    request.session['selected_month'] = month_name  # Store the selected month
    request.session['selected_year'] = year 
    # Render the response
    return render(request, 'audit_salary/pf_category_department_salary.html', {
        'form': form,
        'salary_details': salary_details,
        'category_totals': category_totals,
        'department_totals': department_totals,
        'overall_totals': overall_totals,
        'overall_total_employees': overall_total_employees,
        'selected_month': month_name,  # Pass selected month
    'selected_year': year,  # Pass selected year
        
    })

from .models import PFAbstract

def save_pf_abstract(request):
    if request.method == "POST":
        # Retrieve data from session
        department_totals = request.session.get('department_totals', {})
        selected_month = request.session.get('selected_month', 'Unknown')
        selected_year = request.session.get('selected_year')

        # Handle missing year with a default or error
        if not selected_year:
            messages.error(request, "Year is missing from the session. Please try again.")
            return redirect('pf_category_department_salary')  # Replace with your view name

        try:
            selected_year = int(selected_year)
        except ValueError:
            messages.error(request, "Invalid year format. Please ensure the year is correct.")
            return redirect('pf_category_department_salary')  # Replace with your view name

        # Save or update each department's data
        for department_name, data in department_totals.items():
            obj, created = PFAbstract.objects.update_or_create(
                month=selected_month,
                year=selected_year,
                department_name=department_name,
                defaults={
                    'total_employees': data.get('total_employees', 0),
                    'shift_wages': data.get('shift_wages', 0.0),
                    'total_per_day_salary': data.get('total_per_day_salary', 0.0),
                    'total_rent': data.get('total_rent', 0.0),
                    'total_ot_hours': data.get('total_ot_hours', 0),
                    'total_basic_amount': data.get('total_basic_amount', 0.0),
                    'total_da_amount': data.get('total_da_amount', 0.0),
                    'total_hra_amount': data.get('total_hra_amount', 0.0),
                    'total_pf_amount': data.get('total_pf_amount', 0.0),
                    'total_esi_amount': data.get('total_esi_amount', 0.0),
                    'total_paid_days': data.get('total_paid_days', 0),
                    'advance_total': data.get('advance_total', 0.0),
                    'mess_total': data.get('mess_total', 0.0),
                    'store_total': data.get('store_total', 0.0),
                    'others_total': data.get('others_total', 0.0),
                    'total_deductions': data.get('total_deductions', 0.0),
                    'net_salary_total': data.get('net_salary_total', 0.0),
                    'total_gross_amount': data.get('total_gross_amount', 0.0),
                }
            )
            if created:
                messages.success(request, f"Created record for {department_name} ({selected_month} {selected_year}).")
            else:
                messages.info(request, f"Updated record for {department_name} ({selected_month} {selected_year}).")

        # Redirect to the appropriate page
        messages.success(request, "All department records have been successfully saved.")
        return redirect('pf_category_totals_view')  # Replace with your view name
    else:
        # If accessed via GET, redirect to another page
        return redirect('pf_category_totals_view')  # Replace with your view name



import openpyxl
from openpyxl.styles import Alignment, Font
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Alignment, Font
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse

def pf_export_department_totals_excel(request):
    # Initialize workbook and worksheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Department Totals"
    
    # Set headers
    headers = [
        "Sl.No", "Department", "Total Employees", "Total Paid Days",
        "Total Per Day Salary", "Total Basic", "Total DA", "Total HRA",
        "Total Gross Amount", "Total PF", "Total ESI", "Advance Deductions",
        "Mess Deductions", "Store Deductions", "Other Deductions", "Net Salary"
    ]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Retrieve data from session or fallback
    department_totals = request.session.get('department_totals', {})
    overall_totals = request.session.get('overall_totals', {})
    selected_month =request.session.get('month_name')
    selected_year =request.session.get('year')

    # Write department data
    row_num = 2
    for index, (department_name, department_data) in enumerate(department_totals.items(), start=1):
        sheet.cell(row=row_num, column=1, value=index)  # Sl.No
        sheet.cell(row=row_num, column=2, value=department_name)  # Department Name
        sheet.cell(row=row_num, column=3, value=department_data.get('total_employees', 0))
        sheet.cell(row=row_num, column=4, value=department_data.get('total_paid_days', 0))
        sheet.cell(row=row_num, column=5, value=department_data.get('total_per_day_salary', 0))
        sheet.cell(row=row_num, column=6, value=department_data.get('total_basic_amount', 0))
        sheet.cell(row=row_num, column=7, value=department_data.get('total_da_amount', 0))
        sheet.cell(row=row_num, column=8, value=department_data.get('total_hra_amount', 0))
        sheet.cell(row=row_num, column=9, value=department_data.get('total_gross_amount', 0))
        sheet.cell(row=row_num, column=10, value=department_data.get('total_pf_amount', 0))
        sheet.cell(row=row_num, column=11, value=department_data.get('total_esi_amount', 0))
        sheet.cell(row=row_num, column=12, value=department_data.get('advance_total', 0))
        sheet.cell(row=row_num, column=13, value=department_data.get('mess_total', 0))
        sheet.cell(row=row_num, column=14, value=department_data.get('store_total', 0))
        sheet.cell(row=row_num, column=15, value=department_data.get('others_total', 0))
        sheet.cell(row=row_num, column=16, value=department_data.get('net_salary_total', 0))
        row_num += 1

    # Write overall totals as the last row
    overall_row = [
        "", "Total", overall_totals.get('total_employees', 0),
        overall_totals.get('total_days', 0), overall_totals.get('total_per_day_salary', 0),
        overall_totals.get('total_basic_amount', 0), overall_totals.get('total_da_amount', 0),
        overall_totals.get('total_hra_amount', 0), overall_totals.get('total_gross_amount', 0),
        overall_totals.get('total_pf_amount', 0), overall_totals.get('total_esi_amount', 0),
        overall_totals.get('total_advance_deductions', 0), overall_totals.get('total_mess_deductions', 0),
        overall_totals.get('total_store_deductions', 0), overall_totals.get('total_other_deductions', 0),
        overall_totals.get('net_salary_total', 0)
    ]
    for col_num, value in enumerate(overall_row, 1):
        sheet.cell(row=row_num, column=col_num, value=value)

    filename = f"pf_abstract_{selected_month}_{selected_year}.xlsx"
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename={filename}'
    workbook.save(response)
    return response